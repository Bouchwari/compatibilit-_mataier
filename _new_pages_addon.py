
# ═══════════════════════════════════════════════════════════════════
#  SETTINGS PAGE — إعدادات المؤسسة
# ═══════════════════════════════════════════════════════════════════
class SettingsPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        PageHeader(self, "⚙️  إعدادات المؤسسة",
                   subtitle="معلومات تظهر في الوثائق الرسمية",
                   color="#1A2740").pack(fill="x")

        card = Card(self)
        card.pack(fill="x", padx=20, pady=14)

        tk.Label(card, text="بيانات المؤسسة", font=F(12, True),
                 bg=BG_CARD, fg=ACCENT, anchor="e", padx=14, pady=8).pack(fill="x")

        form = tk.Frame(card, bg=BG_CARD, padx=16); form.pack(fill="x")

        self._vars = {}
        fields = [
            ("اسم المؤسسة",        "school_name"),
            ("الأكاديمية",          "academy"),
            ("المديرية الإقليمية",  "delegation"),
            ("العنوان / المدينة",   "address"),
        ]
        for i, (lbl, key) in enumerate(fields):
            row = tk.Frame(form, bg=BG_CARD); row.pack(fill="x", pady=5)
            tk.Label(row, text=lbl + ":", font=F(11, True),
                     bg=BG_CARD, fg=TEXT_DARK, width=22, anchor="e").pack(side="right")
            v = tk.StringVar()
            self._vars[key] = v
            tk.Entry(row, textvariable=v, font=F(12), justify="right",
                     bg=BG_INPUT, relief="solid", bd=1).pack(
                     side="right", fill="x", expand=True, padx=(0, 8))

        # Logo picker
        logo_row = tk.Frame(form, bg=BG_CARD); logo_row.pack(fill="x", pady=5)
        tk.Label(logo_row, text="شعار المؤسسة:", font=F(11, True),
                 bg=BG_CARD, fg=TEXT_DARK, width=22, anchor="e").pack(side="right")
        self._logo_var = tk.StringVar()
        tk.Entry(logo_row, textvariable=self._logo_var, font=F(10),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1,
                 state="readonly", width=36).pack(side="right", padx=(4, 8))
        ActionBtn(logo_row, "📂 اختيار صورة", command=self._pick_logo,
                  style="primary").pack(side="right")

        # Logo preview label
        self._logo_preview = tk.Label(card, bg=BG_CARD, pady=4)
        self._logo_preview.pack()

        # Buttons
        btn_f = tk.Frame(card, bg=BG_CARD, padx=16, pady=10); btn_f.pack(fill="x")
        ActionBtn(btn_f, "💾  حفظ الإعدادات", command=self._save,
                  style="success").pack(side="right", padx=4)

        tk.Label(self,
                 text="ℹ️  سيظهر الشعار في رأس وثيقة شهادة التسلم",
                 font=F(9), bg=BG_MAIN, fg=TEXT_MID, anchor="e"
                 ).pack(fill="x", padx=20, pady=(4, 0))

        # ── SHARING GUIDE CARD ───────────────────────────────────────
        share_card = Card(self)
        share_card.pack(fill="x", padx=20, pady=12)

        tk.Label(share_card, text="📤  دليل المشاركة مع الزملاء",
                 font=F(12, True), bg=BG_CARD, fg="#0D2137",
                 anchor="e", padx=14, pady=8).pack(fill="x")

        guide_f = tk.Frame(share_card, bg=BG_CARD, padx=20, pady=4)
        guide_f.pack(fill="x")

        # What TO share
        yes_f = tk.Frame(guide_f, bg="#E8F5E9", relief="flat", bd=0, padx=12, pady=8)
        yes_f.pack(fill="x", pady=4)
        tk.Label(yes_f, text="✅  الملفات التي يجب مشاركتها:",
                 font=F(10, True), bg="#E8F5E9", fg="#1B5E20", anchor="e").pack(fill="x")
        tk.Label(yes_f,
                 text="  pages.py  •  main.py  •  database.py  •  theme.py  •  _new_pages_addon.py\n"
                      "  requirements.txt  •  تشغيل.vbs  •  مجلد assets/  •  مجلد templet/  •  مجلد utils/",
                 font=F(9), bg="#E8F5E9", fg="#2E7D32", anchor="e",
                 justify="right").pack(fill="x")

        # What NOT to share
        no_f = tk.Frame(guide_f, bg="#FFEBEE", relief="flat", bd=0, padx=12, pady=8)
        no_f.pack(fill="x", pady=4)
        tk.Label(no_f, text="❌  الملفات التي لا تُشارك (بياناتك الخاصة):",
                 font=F(10, True), bg="#FFEBEE", fg="#B71C1C", anchor="e").pack(fill="x")
        tk.Label(no_f,
                 text="  almadoun.db  (قاعدة البيانات — تحتوي على بياناتك)\n"
                      "  مجلد exports/  •  مجلد __pycache__/  •  مجلد .venv/",
                 font=F(9), bg="#FFEBEE", fg="#C62828", anchor="e",
                 justify="right").pack(fill="x")

        tip_f = tk.Frame(guide_f, bg="#FFF9C4", relief="flat", bd=0, padx=12, pady=8)
        tip_f.pack(fill="x", pady=4)
        tk.Label(tip_f,
                 text="💡  كل مستخدم جديد سيحصل على قاعدة بيانات فارغة تلقائياً عند أول تشغيل، وسيظهر له معالج الإعداد.",
                 font=F(9), bg="#FFF9C4", fg="#5D4037", anchor="e",
                 wraplength=500, justify="right").pack(fill="x")

        # Prepare share folder button
        btn_share_f = tk.Frame(share_card, bg=BG_CARD, padx=14, pady=10)
        btn_share_f.pack(fill="x")

        self._share_status_var = tk.StringVar()
        tk.Label(btn_share_f, textvariable=self._share_status_var,
                 font=F(9), bg=BG_CARD, fg=GREEN, anchor="e").pack(side="right", padx=8)
        ActionBtn(btn_share_f, "📦  إعداد مجلد المشاركة",
                  command=self._prepare_share, style="primary").pack(side="right", padx=4)

        self.refresh()

    def _prepare_share(self):
        try:
            import main as _main
            share_dir, copied = _main._prepare_share_folder()
            import subprocess
            subprocess.Popen(["explorer", share_dir])
            self._share_status_var.set(f"✅  تم إنشاء المجلد: share/  ({len(copied)} ملف)")
        except Exception as e:
            self._share_status_var.set(f"❌  خطأ: {e}")



    def _pick_logo(self):
        path = filedialog.askopenfilename(
            title="اختر صورة الشعار",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp *.gif"),
                       ("All files", "*.*")]
        )
        if path:
            self._logo_var.set(path)
            self._show_preview(path)

    def _show_preview(self, path):
        try:
            from PIL import Image, ImageTk
            img   = Image.open(path).resize((80, 80), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self._logo_preview.config(image=photo, text="")
            self._logo_preview._img = photo   # prevent GC
        except Exception:
            self._logo_preview.config(
                text=f"✅  تم اختيار: {os.path.basename(path)}",
                font=F(9), fg=GREEN)

    def _save(self):
        s = {k: v.get().strip() for k, v in self._vars.items()}
        s["logo_path"] = self._logo_var.get().strip()
        if not s["school_name"]:
            error_dialog(self, "خطأ", "يرجى إدخال اسم المؤسسة على الأقل")
            return
        db.save_settings(
            s["school_name"], s["academy"],
            s["delegation"], s["address"], s["logo_path"]
        )
        info_dialog(self, "تم الحفظ", "✅  تم حفظ إعدادات المؤسسة بنجاح")

    def refresh(self):
        s = db.get_settings()
        for key, v in self._vars.items():
            v.set(s.get(key, ""))
        logo = s.get("logo_path", "")
        self._logo_var.set(logo)
        if logo and os.path.exists(logo):
            self._show_preview(logo)


# ═══════════════════════════════════════════════════════════════════
#  YEAR-END PAGE — إجراءات نهاية السنة
# ═══════════════════════════════════════════════════════════════════
class YearEndPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        PageHeader(self, "📅  إجراءات نهاية / بداية السنة",
                   color="#2E4057").pack(fill="x")

        # Year label input
        top = Card(self); top.pack(fill="x", padx=20, pady=12)
        tk.Label(top, text="السنة الدراسية", font=F(12, True),
                 bg=BG_CARD, fg=ACCENT, anchor="e", padx=14, pady=8).pack(fill="x")

        yr_row = tk.Frame(top, bg=BG_CARD, padx=16, pady=8); yr_row.pack(fill="x")
        tk.Label(yr_row, text="رمز السنة (مثال: 2025-2026):",
                 font=F(11, True), bg=BG_CARD, anchor="e").pack(side="right")
        self._year_var = tk.StringVar()
        now = datetime.now()
        y1  = now.year if now.month >= 9 else now.year - 1
        self._year_var.set(f"{y1}-{y1+1}")
        tk.Entry(yr_row, textvariable=self._year_var, font=F(12),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1,
                 width=16).pack(side="right", padx=8)

        # Close year
        cc = Card(self); cc.pack(fill="x", padx=20, pady=6)
        tk.Label(cc,
                 text="📋  إقفال السنة — تصدير كشف المخزون النهائي",
                 font=F(12, True), bg=BG_CARD, fg="#C55A11",
                 anchor="e", padx=14, pady=8).pack(fill="x")
        tk.Label(cc,
                 text="يولّد ملف Excel يحتوي على الرصيد النهائي لكل مادة، ويحفظ لقطة في قاعدة البيانات.",
                 font=F(10), bg=BG_CARD, fg=TEXT_MID, anchor="e", padx=14).pack(fill="x")
        ActionBtn(cc, "📥  إقفال السنة وتصدير الكشف",
                  command=self._close_year, style="warning").pack(
                  anchor="e", padx=14, pady=8)

        # Open new year
        oc = Card(self); oc.pack(fill="x", padx=20, pady=6)
        tk.Label(oc,
                 text="🆕  فتح سنة جديدة — ترحيل الأرصدة",
                 font=F(12, True), bg=BG_CARD, fg=GREEN,
                 anchor="e", padx=14, pady=8).pack(fill="x")
        tk.Label(oc,
                 text="يُسجّل رصيد كل مادة كـ'دخول' (رصيد مرحّل) في بداية السنة الجديدة.",
                 font=F(10), bg=BG_CARD, fg=TEXT_MID, anchor="e", padx=14).pack(fill="x")
        ActionBtn(oc, "🚀  فتح سنة جديدة وترحيل الأرصدة",
                  command=self._open_new_year, style="success").pack(
                  anchor="e", padx=14, pady=8)

        # History
        tk.Label(self, text="سجل الإقفالات السابقة", font=F(11, True),
                 bg=BG_MAIN, fg=TEXT_DARK, anchor="e", padx=20).pack(
                 fill="x", pady=(8, 2))
        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=20)
        tbl_f.pack(fill="both", expand=True)
        frm, self._hist_tree = make_treeview(
            tbl_f,
            ("year", "closed_at"),
            ("السنة الدراسية", "تاريخ الإقفال"),
            (220, 280), height=6)
        frm.pack(fill="both", expand=True)
        self.refresh()

    # ── Handlers ─────────────────────────────────────────────────────
    def _close_year(self):
        year = self._year_var.get().strip()
        if not year:
            error_dialog(self, "خطأ", "يرجى إدخال رمز السنة الدراسية"); return
        if not confirm_dialog(self, "تأكيد الإقفال",
                f"سيتم حفظ الرصيد النهائي لكل المواد لسنة {year}\n"
                f"وتصدير ملف Excel للكشف النهائي.\n\nهل تريد المتابعة؟"):
            return
        try:
            snapshot = db.close_year(year)
            out_path = self._export_year_excel(year, snapshot)
            import subprocess
            subprocess.Popen(["start", "", out_path], shell=True)
            info_dialog(self, "تم الإقفال",
                f"تم إقفال سنة {year} ✅\n"
                f"{len(snapshot)} مادة مُصنَّفة\n"
                f"📂  {out_path}")
            self.refresh()
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر الإقفال:\n{e}")

    def _open_new_year(self):
        year = self._year_var.get().strip()
        if not year:
            error_dialog(self, "خطأ", "يرجى إدخال رمز السنة الدراسية السابقة"); return
        if not confirm_dialog(self, "⚠️  تأكيد فتح السنة الجديدة",
                f"سيتم إنشاء حركات دخول (رصيد مرحّل من {year})\n"
                f"لكل مادة رصيدها أكبر من صفر.\n\n"
                f"تأكد من إقفال السنة السابقة أولاً!\nهل تريد المتابعة؟"):
            return
        try:
            count = db.open_new_year(year)
            info_dialog(self, "تم الترحيل",
                f"✅  تم ترحيل الأرصدة\n"
                f"📦  {count} مادة مُرحَّلة كـ'دخول' في بداية السنة الجديدة")
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر الترحيل:\n{e}")

    def _export_year_excel(self, year_label, snapshot):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الكشف النهائي"
        ws.sheet_view.rightToLeft = True

        # Styles
        hdr_font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="1A2740")
        hdr_align = Alignment(horizontal="center", vertical="center")
        c_align   = Alignment(horizontal="right", vertical="center")
        thin      = Side(style="thin", color="D0D7E3")
        bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
        cat_fill  = PatternFill("solid", fgColor="E8EDF5")
        red_fill  = PatternFill("solid", fgColor="FFE5E5")
        amber_fill= PatternFill("solid", fgColor="FFF8E1")

        settings = db.get_settings()
        school   = settings.get("school_name", "المؤسسة") or "المؤسسة"

        # Title
        ws.merge_cells("A1:F1")
        tc = ws["A1"]
        tc.value = f"كشف المخزون النهائي — {school} — {year_label}"
        tc.font  = Font(name="Calibri", bold=True, size=14)
        tc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        # Headers
        hdrs = ["رقم", "اسم المادة", "الفئة", "الرصيد النهائي", "الوحدة", "ملاحظات"]
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = bdr
        ws.row_dimensions[2].height = 22

        prev_cat = None; ri = 3; idx = 0
        for item in snapshot:
            cat = item.get("category", "")
            if cat != prev_cat:
                ws.merge_cells(f"A{ri}:F{ri}")
                cat_c = ws.cell(row=ri, column=1, value=f"▶  {cat}")
                cat_c.font = Font(bold=True, size=11)
                cat_c.fill = cat_fill; cat_c.alignment = c_align
                ws.row_dimensions[ri].height = 18
                ri += 1; prev_cat = cat
            idx += 1
            bal = item.get("balance", 0)
            fill = red_fill if bal <= 0 else (amber_fill if bal <= 5 else None)
            row_data = [idx, item.get("name",""), cat, bal, item.get("unit",""), ""]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = c_align; c.border = bdr
                if fill: c.fill = fill
            ws.row_dimensions[ri].height = 16; ri += 1

        for ci, w in enumerate([8, 42, 22, 16, 14, 20], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        from datetime import datetime as _dt
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
        os.makedirs(exports_dir, exist_ok=True)
        fname = f"إقفال_السنة_{year_label}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        out   = os.path.join(exports_dir, fname)
        wb.save(out)
        return out

    def refresh(self):
        for row in self._hist_tree.get_children():
            self._hist_tree.delete(row)
        for i, r in enumerate(db.get_year_closings()):
            tag = "even" if i % 2 == 0 else "odd"
            self._hist_tree.insert("", "end",
                values=(r["year_label"], r["closed_at"]), tags=(tag,))
