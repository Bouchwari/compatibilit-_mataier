"""
pages.py — All UI pages for the app
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import os
import sys

# Fix matplotlib backend BEFORE any other matplotlib import
import matplotlib
matplotlib.use("TkAgg")

import database as db
from theme import *

# Document export utilities
try:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from utils import doc_export
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

MONTHS_AR = ["يناير","فبراير","مارس","أبريل","ماي","يونيو",
             "يوليوز","غشت","شتنبر","أكتوبر","نونبر","دجنبر"]

# ═══════════════════════════════════════════════════════════════════
#  HOME / DASHBOARD
# ═══════════════════════════════════════════════════════════════════
class HomePage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        # Header
        PageHeader(self,
            "🏠  الثانوية الإعدادية ألمدون — نظام إدارة المستهلكات",
            subtitle="اختر قسماً من القائمة الجانبية للبدء",
            color=BG_DARK
        ).pack(fill="x")

        # KPI row
        kpi_frame = tk.Frame(self, bg=BG_MAIN, padx=20, pady=15)
        kpi_frame.pack(fill="x")

        self.kpi_cards = {}
        kpis = [
            ("🗂️","إجمالي الأصناف","total",  ACCENT),
            ("✅","متوفر",           "available", GREEN),
            ("⚠️","مخزون منخفض",    "low",    ORANGE),
            ("🚨","نفد المخزون",    "out",    RED),
            ("📋","عدد العمليات",   "operations","#5C3E8F"),
        ]
        for i,(icon,label,key,color) in enumerate(kpis):
            card = KPICard(kpi_frame, icon, label, "—", color=color)
            card.grid(row=0, column=i, padx=8, pady=4, sticky="nsew")
            kpi_frame.columnconfigure(i, weight=1)
            self.kpi_cards[key] = card

        # Recent operations
        tk.Label(self, text="آخر العمليات المسجلة",
            font=F(12,True), bg=BG_MAIN, fg=TEXT_DARK,
            anchor="e", padx=20, pady=8).pack(fill="x")

        tbl_frame = tk.Frame(self, bg=BG_MAIN, padx=20)
        tbl_frame.pack(fill="both", expand=True, pady=(0,10))

        cols  = ("date","item","category","type","qty","beneficiary")
        hdrs  = ("التاريخ","المادة","الفئة","النوع","الكمية","المستفيد / المورد")
        widths= (100, 220, 130, 80, 70, 180)
        frame, self.recent_tree = make_treeview(tbl_frame, cols, hdrs, widths, height=14)
        frame.pack(fill="both", expand=True)

        self.refresh()

    def refresh(self):
        # KPIs
        kpi = db.get_kpi_summary()
        for key, card in self.kpi_cards.items():
            card.update_value(kpi.get(key, 0))

        # Recent
        for row in self.recent_tree.get_children():
            self.recent_tree.delete(row)
        movements = db.get_movements(limit=50)
        for i, m in enumerate(movements):
            tag = "in" if m["type"]=="دخول" else "odd"
            self.recent_tree.insert("", "end", values=(
                m["date"] or "",
                m["item_name"],
                m["category"] or "",
                m["type"],
                int(m["quantity"]),
                m["beneficiary"] or ""
            ), tags=(tag if m["type"]=="دخول" else ("even" if i%2==0 else "odd"),))


# ═══════════════════════════════════════════════════════════════════
#  DASHBOARD / INVENTORY STATUS
# ═══════════════════════════════════════════════════════════════════
class DashboardPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        PageHeader(self, "📊  لوحة التحكم — حالة المخزون", color=ACCENT).pack(fill="x")

        # Filter bar
        filter_f = tk.Frame(self, bg=BG_MAIN, padx=16, pady=8)
        filter_f.pack(fill="x")
        tk.Label(filter_f, text="تصفية:", font=F(10,True), bg=BG_MAIN).pack(side="right")
        self.filter_var = tk.StringVar(value="الكل")
        cats = ["الكل"] + list(db.CATEGORIES.keys())
        cb = ttk.Combobox(filter_f, textvariable=self.filter_var,
                          values=cats, width=20, state="readonly", font=F(11))
        cb.pack(side="right", padx=6)
        cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        self.show_var = tk.StringVar(value="الكل")
        show_cb = ttk.Combobox(filter_f, textvariable=self.show_var,
            values=["الكل","🟢 متوفر","🟡 منخفض","🔴 نفد"], width=14,
            state="readonly", font=F(11))
        show_cb.pack(side="right", padx=6)
        show_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        # Search
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *a: self.refresh())
        tk.Entry(filter_f, textvariable=self.search_var, width=22, font=F(11),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1).pack(side="right", padx=6)
        tk.Label(filter_f, text="🔍", font=F(12), bg=BG_MAIN).pack(side="right")

        ActionBtn(filter_f, "🔄 تحديث", command=self.refresh, style="neutral").pack(side="left", padx=4)
        ActionBtn(filter_f, "📊 تصدير Excel", command=self._export_excel, style="primary").pack(side="left", padx=4)
        ActionBtn(filter_f, "📋 قالب الجرد", command=self._export_template, style="success").pack(side="left", padx=4)
        ActionBtn(filter_f, "📤 استيراد الجرد", command=self._import_inventory, style="warning").pack(side="left", padx=4)

        # Treeview
        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=16)
        tbl_f.pack(fill="both", expand=True, pady=(0,10))

        cols  = ("category","item","unit","in","out","balance","status","last_date")
        hdrs  = ("الفئة","التسمية","الوحدة","الدخول","الخروج","الرصيد","الحالة","آخر عملية")
        widths= (110, 220, 80, 80, 80, 90, 120, 100)
        frame, self.tree = make_treeview(tbl_f, cols, hdrs, widths, height=22)
        frame.pack(fill="both", expand=True)

        self.refresh()

    def _export_excel(self):
        try:
            path = doc_export.export_inventory_excel(db.get_dashboard_data())
            import subprocess
            subprocess.Popen(["explorer", "/select,", path])
            info_dialog(self, "تصدير Excel",
                f"تم تصدير كشف المخزون إلى:\n{path}")
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر التصدير:\n{e}")

    def _export_template(self):
        """Export a blank inventory Excel template pre-filled with all items."""
        try:
            data = db.get_dashboard_data()
            path = doc_export.export_inventory_template_excel(data)
            import subprocess
            subprocess.Popen(["explorer", "/select,", path])
            info_dialog(self, "قالب الجرد",
                f"تم تصدير قالب الجرد إلى:\n{path}\n\n"
                "خطوات الاستخدام:\n"
                "1️⃣  افتح الملف في Excel\n"
                "2️⃣  أدخل الكميات في العمود E (المصفّر بالأصفر)\n"
                "3️⃣  احفظ واستخدم 'استيراد الجرد' لاستيراده")
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر التصدير:\n{e}")

    def _import_inventory(self):
        """Import a filled inventory Excel template and create stock movements."""
        from tkinter.filedialog import askopenfilename
        import openpyxl

        path = askopenfilename(
            title="اختر ملف الجرد",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return

        # ── Step 1: Preview what will be imported ─────────────────
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            preview_rows = []
            for row in ws.iter_rows(min_row=4, values_only=True):
                name = str(row[1] or "").strip()
                if not name:
                    continue
                try:
                    qty = float(row[4] or 0)
                except Exception:
                    qty = 0
                if qty > 0:
                    preview_rows.append(f"  • {name}  →  +{int(qty)}")
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر قراءة الملف:\n{e}")
            return

        if not preview_rows:
            error_dialog(self, "لا يوجد ما يُستورد",
                "لم يتم العثور على كميات في العمود E.\n\n"
                "تأكد أنك أدخلت الأرقام في العمود 'الكمية المستلمة' (E)\n"
                "المُعلَّم باللون الأصفر في قالب الجرد.")
            return

        preview_text = "\n".join(preview_rows[:15])
        if len(preview_rows) > 15:
            preview_text += f"\n  ... و{len(preview_rows)-15} مادة أخرى"

        confirm = messagebox.askyesno(
            "تأكيد الاستيراد",
            f"سيتم تسجيل {len(preview_rows)} عملية دخول:\n\n"
            f"{preview_text}\n\n"
            "هل تريد المتابعة؟",
            icon="question"
        )
        if not confirm:
            return

        # ── Step 2: Actually import ────────────────────────────────
        try:
            imported, skipped, errors = doc_export.import_inventory_from_excel(path)
            msg = f"✅  تم استيراد {imported} عملية دخول بنجاح\n"
            if skipped:
                msg += f"⏭️  تم تخطّي {len(skipped)} مادة (كمية = 0)\n"
            if errors:
                msg += f"⚠️  {len(errors)} خطأ:\n" + "\n".join(errors[:5])
            info_dialog(self, "اكتمل الاستيراد", msg)
            # ── Step 3: Refresh dashboard so changes are visible ───
            self.refresh()
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر الاستيراد:\n{e}")


    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        cat_filter  = self.filter_var.get()
        show_filter = self.show_var.get()
        search      = self.search_var.get().strip().lower()

        data = db.get_dashboard_data()
        i = 0
        for r in data:
            if cat_filter != "الكل" and r["category"] != cat_filter:
                continue
            if search and search not in r["name"].lower():
                continue

            balance = r["balance"]
            if balance <= 0:
                status = "🔴  نفد المخزون"; tag = "out"
            elif balance <= 5:
                status = "🟡  منخفض"; tag = "low"
            else:
                status = "🟢  متوفر"; tag = "even" if i%2==0 else "odd"

            if show_filter == "🟢 متوفر" and "متوفر" not in status: continue
            if show_filter == "🟡 منخفض" and "منخفض" not in status: continue
            if show_filter == "🔴 نفد"   and "نفد"    not in status: continue

            self.tree.insert("", "end", values=(
                r["category"], r["name"], r["unit"] or "",
                int(r["total_in"]), int(r["total_out"]),
                int(balance), status,
                r["last_date"] or "—"
            ), tags=(tag,))
            i += 1


# ═══════════════════════════════════════════════════════════════════
#  MOVEMENTS PAGE (per category)
# ═══════════════════════════════════════════════════════════════════
class MovementsPage(tk.Frame):
    def __init__(self, parent, category, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self.category = category
        self.color = CAT_COLORS.get(category, ACCENT)
        self._build()

    def _build(self):
        color = self.color
        PageHeader(self, f"📁  {self.category} — سجل الحركة",
                   subtitle="دخول وخروج المواد مع الرصيد التلقائي",
                   color=color).pack(fill="x")

        # ── ADD FORM ───────────────────────────────────────────────
        form_card = Card(self)
        form_card.pack(fill="x", padx=16, pady=(10,0))

        tk.Label(form_card, text="➕  تسجيل عملية جديدة",
            font=F(11,True), bg=BG_CARD, fg=color, anchor="e",
            padx=14, pady=8).pack(fill="x")

        fields = tk.Frame(form_card, bg=BG_CARD, padx=14, pady=4)
        fields.pack(fill="x")

        # Row 1
        r1 = tk.Frame(fields, bg=BG_CARD); r1.pack(fill="x", pady=2)

        # Date
        tk.Label(r1, text="التاريخ :", font=F(10,True), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0,4))
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(r1, textvariable=self.date_var, width=14, font=F(11),
                 justify="center", bg=BG_INPUT, relief="solid", bd=1).pack(side="right", padx=(0,12))

        # Type
        tk.Label(r1, text="النوع :", font=F(10,True), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0,4))
        self.type_var = tk.StringVar(value="دخول")
        type_cb = ttk.Combobox(r1, textvariable=self.type_var,
                     values=["دخول","خروج"], width=10, state="readonly", font=F(11))
        type_cb.pack(side="right", padx=(0,12))
        type_cb.bind("<<ComboboxSelected>>", self._on_type_change)

        # Quantity
        tk.Label(r1, text="الكمية :", font=F(10,True), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0,4))
        self.qty_var = tk.StringVar()
        tk.Entry(r1, textvariable=self.qty_var, width=10, font=F(11),
                 justify="center", bg=BG_INPUT, relief="solid", bd=1).pack(side="right", padx=(0,12))

        # Row 2
        r2 = tk.Frame(fields, bg=BG_CARD); r2.pack(fill="x", pady=2)

        # Item
        tk.Label(r2, text="المادة :", font=F(10,True), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0,4))
        items = [row["name"] for row in db.get_items_by_category(self.category)]
        self.item_var = tk.StringVar()
        self.item_cb = ttk.Combobox(r2, textvariable=self.item_var,
                    values=items, width=32, font=F(11))
        self.item_cb.pack(side="right", padx=(0,12))
        self.item_cb.bind("<<ComboboxSelected>>", self._on_item_select)

        # Unit (auto)
        tk.Label(r2, text="الوحدة :", font=F(10,True), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0,4))
        self.unit_var = tk.StringVar()
        tk.Entry(r2, textvariable=self.unit_var, width=12, font=F(11),
                 justify="center", bg=BG_ROW_ALT, relief="solid", bd=1,
                 state="readonly").pack(side="right", padx=(0,12))

        # Row 3 — المورد (for دخول) or المستفيد (for خروج), dynamic
        r3 = tk.Frame(fields, bg=BG_CARD); r3.pack(fill="x", pady=2)
        self.benef_label = tk.Label(r3, text="المورد :",
            font=F(10,True), bg=BG_CARD, fg=TEXT_DARK)
        self.benef_label.pack(side="right", padx=(0,4))
        self._staff_names    = db.get_staff_names()
        self._supplier_names = db.get_supplier_names()
        self.benef_var = tk.StringVar()
        self.benef_cb = ttk.Combobox(r3, textvariable=self.benef_var,
                     values=self._supplier_names, width=28, font=F(11))
        self.benef_cb.pack(side="right", padx=(0,12))
        tk.Label(r3, text="ملاحظات :", font=F(10,True), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0,4))
        self.notes_var = tk.StringVar()
        tk.Entry(r3, textvariable=self.notes_var, width=30, font=F(11),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1).pack(side="right", padx=(0,12))

        # Buttons
        btn_f = tk.Frame(form_card, bg=BG_CARD, padx=14, pady=10)
        btn_f.pack(fill="x")
        ActionBtn(btn_f, "✅  حفظ العملية", command=self._save, style="success").pack(side="right", padx=4)
        ActionBtn(btn_f, "🗑️  مسح الحقول", command=self._clear, style="neutral").pack(side="right", padx=4)

        # ── BALANCE INDICATOR ────────────────────────────────────────
        bal_f = tk.Frame(self, bg=BG_MAIN, padx=16, pady=4)
        bal_f.pack(fill="x")
        self.bal_label = tk.Label(bal_f, text="", font=F(12,True),
                                   bg=BG_MAIN, fg=GREEN, anchor="e")
        self.bal_label.pack(side="right")

        # ── MOVEMENTS TABLE ──────────────────────────────────────────
        tk.Label(self, text="سجل العمليات", font=F(11,True),
                 bg=BG_MAIN, fg=TEXT_DARK, anchor="e", padx=16).pack(fill="x")

        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=16)
        tbl_f.pack(fill="both", expand=True, pady=(0,8))

        cols  = ("id","date","item","type","qty","balance","beneficiary","notes")
        hdrs  = ("رقم","التاريخ","المادة","النوع","الكمية","الرصيد","المستفيد","ملاحظات")
        widths= (50, 100, 200, 70, 70, 80, 160, 160)
        frame, self.tree = make_treeview(tbl_f, cols, hdrs, widths, height=14)
        frame.pack(fill="both", expand=True)

        btn_f2 = tk.Frame(self, bg=BG_MAIN, padx=16, pady=4)
        btn_f2.pack(fill="x")
        ActionBtn(btn_f2, "🗑️  حذف المحدد", command=self._delete, style="danger").pack(side="right", padx=4)
        ActionBtn(btn_f2, "🔄  تحديث", command=self.refresh, style="neutral").pack(side="right", padx=4)

        self.refresh()

    def _on_type_change(self, e=None):
        if self.type_var.get() == "دخول":
            self.benef_label.config(text="المورد :")
            self.benef_cb["values"] = self._supplier_names
        else:
            self.benef_label.config(text="المستفيد :")
            self.benef_cb["values"] = self._staff_names
        self.benef_var.set("")

    def _on_item_select(self, e):
        unit = db.get_item_unit(self.item_var.get())
        self.unit_var.set(unit)
        # Show current balance
        in_, out_, bal = db.get_item_balance(self.item_var.get())
        color = GREEN if bal > 5 else (ORANGE if bal > 0 else RED)
        self.bal_label.config(
            text=f"الرصيد الحالي:  {int(in_)} دخول  —  {int(out_)} خروج  =  {int(bal)}",
            fg=color)

    def _save(self):
        item = self.item_var.get().strip()
        qty_s = self.qty_var.get().strip()
        date = self.date_var.get().strip()
        typ = self.type_var.get()

        if not item:
            error_dialog(self, "خطأ", "يرجى اختيار المادة"); return
        if not qty_s or not qty_s.replace(".","").isdigit():
            error_dialog(self, "خطأ", "يرجى إدخال كمية صحيحة"); return

        qty = float(qty_s)
        if qty <= 0:
            error_dialog(self, "خطأ", "الكمية يجب أن تكون أكبر من صفر"); return

        # Check balance for خروج
        if typ == "خروج":
            _,_, bal = db.get_item_balance(item)
            if qty > bal:
                error_dialog(self, "خطأ",
                    f"الكمية المطلوبة ({int(qty)}) أكبر من الرصيد المتاح ({int(bal)})"); return

        db.add_movement(item, self.category, date, typ, qty,
                        self.benef_var.get(), self.notes_var.get())
        self._clear()
        self.refresh()
        info_dialog(self, "تم", f"تم تسجيل العملية بنجاح ✅")

    def _clear(self):
        self.item_var.set(""); self.qty_var.set("")
        self.unit_var.set(""); self.benef_var.set(""); self.notes_var.set("")
        self.date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.bal_label.config(text="")
        self._on_type_change()

    def _delete(self):
        sel = self.tree.selection()
        if not sel: return
        row_id = self.tree.item(sel[0])["values"][0]
        if confirm_dialog(self, "تأكيد", "هل تريد حذف هذه العملية؟"):
            db.delete_movement(row_id)
            self.refresh()


    def refresh(self):
        # Reload dropdown data in case staff/suppliers/items changed
        self._staff_names    = db.get_staff_names()
        self._supplier_names = db.get_supplier_names()
        self._on_type_change()
        items = [row["name"] for row in db.get_items_by_category(self.category)]
        self.item_cb["values"] = items

        for row in self.tree.get_children():
            self.tree.delete(row)

        movements = db.get_movements(category=self.category)
        # Build running balance per item
        balances = {}
        ordered = list(reversed(movements))  # oldest first for balance calc
        running = {}
        balance_map = {}
        for m in ordered:
            n = m["item_name"]
            running.setdefault(n, 0)
            if m["type"] == "دخول":
                running[n] += m["quantity"]
            else:
                running[n] -= m["quantity"]
            balance_map[m["id"]] = running[n]

        for i, m in enumerate(movements):
            bal = balance_map.get(m["id"], 0)
            if m["type"] == "دخول":
                tag = "in"
            elif bal <= 0:
                tag = "out"
            elif bal <= 5:
                tag = "low"
            else:
                tag = "even" if i%2==0 else "odd"

            self.tree.insert("", "end", values=(
                m["id"], m["date"] or "", m["item_name"],
                m["type"], int(m["quantity"]), int(bal),
                m["beneficiary"] or "", m["notes"] or ""
            ), tags=(tag,))


# ═══════════════════════════════════════════════════════════════════
#  STAFF PAGE
# ═══════════════════════════════════════════════════════════════════
class StaffPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._selected_id = None
        self._build()

    def _build(self):
        PageHeader(self, "👥  سجل الأساتذة والموظفين",
                   subtitle="الأسماء تظهر في شهادات التسليم", color="#1B6CA8").pack(fill="x")

        # ── FORM ────────────────────────────────────────────────────
        form = Card(self)
        form.pack(fill="x", padx=16, pady=10)
        tk.Label(form, text="إضافة / تعديل موظف", font=F(11,True),
                 bg=BG_CARD, fg=ACCENT, anchor="e", padx=14, pady=8).pack(fill="x")

        fields_f = tk.Frame(form, bg=BG_CARD, padx=14); fields_f.pack(fill="x")
        self.vars = {}
        labels = ["الاسم الكامل","الصفة / المهنة","المادة المُدرَّسة","رقم التأجير","البريد الإلكتروني","ملاحظات"]
        keys   = ["name","position","subject","hire_number","email","notes"]
        for row_i in range(0, len(labels), 3):
            r = tk.Frame(fields_f, bg=BG_CARD); r.pack(fill="x", pady=3)
            for col_i in range(3):
                idx = row_i + col_i
                if idx >= len(labels): break
                lbl, key = labels[idx], keys[idx]
                c = tk.Frame(r, bg=BG_CARD); c.pack(side="right", padx=8, expand=True, fill="x")
                tk.Label(c, text=lbl+":", font=F(10,True), bg=BG_CARD, anchor="e").pack(anchor="e")
                v = tk.StringVar()
                self.vars[key] = v
                tk.Entry(c, textvariable=v, font=F(11), justify="right",
                         bg=BG_INPUT, relief="solid", bd=1, width=22).pack(fill="x")

        btn_f = tk.Frame(form, bg=BG_CARD, padx=14, pady=10); btn_f.pack(fill="x")
        ActionBtn(btn_f, "✅  حفظ", command=self._save, style="success").pack(side="right", padx=4)
        ActionBtn(btn_f, "✏️  تعديل المحدد", command=self._load_selected, style="primary").pack(side="right", padx=4)
        ActionBtn(btn_f, "🗑️  حذف المحدد", command=self._delete, style="danger").pack(side="right", padx=4)
        ActionBtn(btn_f, "🔄  مسح", command=self._clear, style="neutral").pack(side="right", padx=4)

        # ── TABLE ────────────────────────────────────────────────────
        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=16); tbl_f.pack(fill="both", expand=True, pady=(0,10))
        cols  = ("id","name","position","subject","hire_number","notes")
        hdrs  = ("رقم","الاسم","الصفة","المادة","رقم التأجير","ملاحظات")
        widths= (40, 180, 120, 150, 110, 170)
        frame, self.tree = make_treeview(tbl_f, cols, hdrs, widths, height=15)
        frame.pack(fill="both", expand=True)
        self.refresh()

    def _save(self):
        name = self.vars["name"].get().strip()
        if not name:
            error_dialog(self,"خطأ","يرجى إدخال الاسم"); return
        pos   = self.vars["position"].get()
        subj  = self.vars["subject"].get()
        hire  = self.vars["hire_number"].get()
        email = self.vars["email"].get()
        note  = self.vars["notes"].get()
        if self._selected_id:
            db.update_staff(self._selected_id, name, pos, subj, hire, email, note)
        else:
            db.add_staff(name, pos, subj, hire, email, note)
        self._clear(); self.refresh()

    def _load_selected(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])["values"]
        self._selected_id = vals[0]
        keys = ["name","position","subject","hire_number","notes"]
        # vals: id, name, position, subject, hire_number, notes
        for i, k in enumerate(keys):
            self.vars[k].set(vals[i+1] if i+1 < len(vals) else "")

    def _delete(self):
        sel = self.tree.selection()
        if not sel: return
        sid = self.tree.item(sel[0])["values"][0]
        if confirm_dialog(self,"تأكيد","حذف هذا الموظف؟"):
            db.delete_staff(sid); self.refresh()

    def _clear(self):
        for v in self.vars.values(): v.set("")
        self._selected_id = None

    def refresh(self):
        for row in self.tree.get_children(): self.tree.delete(row)
        for i,s in enumerate(db.get_all_staff()):
            tag = "even" if i%2==0 else "odd"
            self.tree.insert("","end",values=(
                s["id"], s["name"], s["position"] or "",
                s["subject"] or "", s["hire_number"] or "",
                s["notes"] or ""
            ), tags=(tag,))


# ═══════════════════════════════════════════════════════════════════
#  SUPPLIERS PAGE
# ═══════════════════════════════════════════════════════════════════

class SuppliersPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._selected_id = None
        self._build()

    def _build(self):
        PageHeader(self, "🏪  سجل الموردين", color="#1B998B").pack(fill="x")

        form = Card(self)
        form.pack(fill="x", padx=16, pady=10)
        tk.Label(form, text="إضافة / تعديل مورد", font=F(11,True),
                 bg=BG_CARD, fg=ACCENT2, anchor="e", padx=14, pady=8).pack(fill="x")

        fields_f = tk.Frame(form, bg=BG_CARD, padx=14); fields_f.pack(fill="x")
        self.vars = {}
        labels = ["اسم المورد","نوع التوريد","رقم الهاتف","العنوان","رقم الفاتورة الأخيرة","تاريخ آخر توريد","ملاحظات"]
        keys   = ["name","supply_type","phone","address","last_invoice","last_date","notes"]

        for row_i in range(0, len(labels), 4):
            r = tk.Frame(fields_f, bg=BG_CARD); r.pack(fill="x", pady=3)
            for col_i in range(4):
                idx = row_i + col_i
                if idx >= len(labels): break
                c = tk.Frame(r, bg=BG_CARD); c.pack(side="right", padx=6, expand=True, fill="x")
                tk.Label(c, text=labels[idx]+":", font=F(10,True), bg=BG_CARD, anchor="e").pack(anchor="e")
                v = tk.StringVar()
                self.vars[keys[idx]] = v
                tk.Entry(c, textvariable=v, font=F(11), justify="right",
                         bg=BG_INPUT, relief="solid", bd=1, width=18).pack(fill="x")

        btn_f = tk.Frame(form, bg=BG_CARD, padx=14, pady=10); btn_f.pack(fill="x")
        ActionBtn(btn_f,"✅  حفظ",command=self._save,style="success").pack(side="right",padx=4)
        ActionBtn(btn_f,"✏️  تعديل",command=self._load,style="primary").pack(side="right",padx=4)
        ActionBtn(btn_f,"🗑️  حذف",command=self._delete,style="danger").pack(side="right",padx=4)
        ActionBtn(btn_f,"🔄  مسح",command=self._clear,style="neutral").pack(side="right",padx=4)

        tbl_f = tk.Frame(self,bg=BG_MAIN,padx=16); tbl_f.pack(fill="both",expand=True,pady=(0,10))
        cols  = ("id","name","type","phone","address","invoice","date","notes")
        hdrs  = ("رقم","المورد","نوع التوريد","الهاتف","العنوان","الفاتورة","آخر توريد","ملاحظات")
        widths= (40,160,120,100,160,120,100,140)
        frame, self.tree = make_treeview(tbl_f,cols,hdrs,widths,height=15)
        frame.pack(fill="both",expand=True)
        self.refresh()

    def _save(self):
        name = self.vars["name"].get().strip()
        if not name: error_dialog(self,"خطأ","يرجى إدخال اسم المورد"); return
        args = [self.vars[k].get() for k in ["name","supply_type","phone","address","last_invoice","last_date","notes"]]
        if self._selected_id: db.update_supplier(self._selected_id, *args)
        else: db.add_supplier(*args)
        self._clear(); self.refresh()

    def _load(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])["values"]
        self._selected_id = vals[0]
        keys = ["name","supply_type","phone","address","last_invoice","last_date","notes"]
        for i,k in enumerate(keys): self.vars[k].set(vals[i+1])

    def _delete(self):
        sel = self.tree.selection()
        if not sel: return
        sid = self.tree.item(sel[0])["values"][0]
        if confirm_dialog(self,"تأكيد","حذف هذا المورد؟"):
            db.delete_supplier(sid); self.refresh()

    def _clear(self):
        for v in self.vars.values(): v.set("")
        self._selected_id = None

    def refresh(self):
        for row in self.tree.get_children(): self.tree.delete(row)
        for i,s in enumerate(db.get_all_suppliers()):
            tag = "even" if i%2==0 else "odd"
            self.tree.insert("","end",values=(
                s["id"],s["name"],s["supply_type"] or "",s["phone"] or "",
                s["address"] or "",s["last_invoice"] or "",s["last_date"] or "",
                s["notes"] or ""
            ), tags=(tag,))


# ═══════════════════════════════════════════════════════════════════
#  CERTIFICATE PAGE
# ═══════════════════════════════════════════════════════════════════
class CertificatePage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self.item_rows = []
        self._build()

    def _build(self):
        PageHeader(self, "📋  شهادة التسلم",
                   subtitle="الثانوية الإعدادية ألمدون — وثيقة إدارية رسمية",
                   color=GOLD).pack(fill="x")

        scroll_outer, scroll_inner = scrollable_frame(self, bg=BG_MAIN)
        scroll_outer.pack(fill="both", expand=True, padx=0, pady=0)

        # ── CERT HEADER ─────────────────────────────────────────────
        header_card = Card(scroll_inner)
        header_card.pack(fill="x", padx=16, pady=10)

        # Official header text
        for txt, sz, bold in [
            ("المملكة المغربية — وزارة التربية الوطنية والتعليم الأولي والرياضة", 10, True),
            ("الأكاديمية الجهوية للتربية والتكوين لجهة درعة تافيلالت", 11, True),
            ("المديرية الإقليمية بتنغير", 11, True),
            ("الثانوية الإعدادية ألمدون — جماعة اغيل نومكون", 12, True),
        ]:
            tk.Label(header_card, text=txt, font=F(sz,bold),
                     bg=BG_CARD, fg=TEXT_DARK, anchor="center").pack(fill="x", pady=1)

        tk.Frame(header_card, bg=BG_DARK, height=2).pack(fill="x", padx=10, pady=6)

        # Cert number + Date row
        top_row = tk.Frame(header_card, bg=BG_CARD); top_row.pack(fill="x", padx=14)
        self.cert_num_var = tk.StringVar(value=str(db.get_next_cert_number()))
        tk.Label(top_row, text="رقم الشهادة:", font=F(10,True), bg=BG_CARD).pack(side="right", padx=4)
        tk.Entry(top_row, textvariable=self.cert_num_var, width=8, font=F(11),
                 justify="center", bg=GOLD_LT, relief="solid", bd=1).pack(side="right", padx=(0,20))
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Label(top_row, text="التاريخ:", font=F(10,True), bg=BG_CARD).pack(side="right", padx=4)
        tk.Entry(top_row, textvariable=self.date_var, width=14, font=F(11),
                 justify="center", bg=GOLD_LT, relief="solid", bd=1).pack(side="right")

        tk.Label(header_card, text="شهادة التسلم",
            font=F(16,True), bg=BG_CARD, fg=BG_DARK,
            anchor="center", pady=10).pack(fill="x")

        # ── RECIPIENT ───────────────────────────────────────────────
        recip_card = Card(scroll_inner)
        recip_card.pack(fill="x", padx=16, pady=(0,8))

        tk.Label(recip_card, text="أنا الموقع(ة) أسفله السيد(ة):",
            font=F(12,True), bg=BG_CARD, fg=TEXT_DARK,
            anchor="e", padx=14, pady=8).pack(fill="x")

        row1 = tk.Frame(recip_card, bg=BG_CARD, padx=14); row1.pack(fill="x")
        tk.Label(row1, text="الاسم والنسب:", font=F(11,True), bg=BG_CARD).pack(side="right", padx=4)

        staff_names = db.get_staff_names()
        self.recip_var = tk.StringVar()
        recip_cb = ttk.Combobox(row1, textvariable=self.recip_var,
                    values=staff_names, width=30, font=F(11))
        recip_cb.pack(side="right", padx=(0,20))
        recip_cb.bind("<<ComboboxSelected>>", self._on_recip_select)

        tk.Label(row1, text="الصفة:", font=F(11,True), bg=BG_CARD).pack(side="right", padx=4)
        self.pos_var = tk.StringVar()
        tk.Entry(row1, textvariable=self.pos_var, width=22, font=F(11),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1).pack(side="right")

        tk.Label(recip_card,
            text="أشهد أني تسلمت من السيد: مسير المصالح المادية والمالية — المواد المبينة في الجدول أسفله:",
            font=F(11), bg=BG_CARD, fg=TEXT_DARK, anchor="e",
            padx=14, pady=6).pack(fill="x")

        # ── ITEMS TABLE ─────────────────────────────────────────────
        items_card = Card(scroll_inner)
        items_card.pack(fill="x", padx=16, pady=(0,8))

        # Table header
        hdr_row = tk.Frame(items_card, bg=BG_DARK); hdr_row.pack(fill="x")
        for txt, w in [("ر.ت",4),("تعيين المادة",30),("الوحدة",10),("العدد المسلَّم",10)]:
            tk.Label(hdr_row, text=txt, font=F(10,True), bg=BG_DARK, fg=TEXT_WHITE,
                     width=w, anchor="center", pady=7).pack(side="right")

        self.item_rows = []
        all_items = [row["name"] for row in db.get_all_items()]

        for i in range(8):
            row_f = tk.Frame(items_card,
                             bg=BG_CARD if i%2==0 else BG_ROW_ALT)
            row_f.pack(fill="x")

            tk.Label(row_f, text=str(i+1), font=F(11,True),
                     bg=row_f["bg"], fg=TEXT_MID, width=4,
                     anchor="center", pady=5).pack(side="right")

            item_var = tk.StringVar()
            item_cb2 = ttk.Combobox(row_f, textvariable=item_var,
                        values=all_items, width=30, font=F(11))
            item_cb2.pack(side="right", padx=4, pady=3)

            unit_var = tk.StringVar()
            unit_e = tk.Entry(row_f, textvariable=unit_var, width=12, font=F(10),
                               justify="center", bg=row_f["bg"], relief="flat", state="readonly")
            unit_e.pack(side="right", padx=4)

            qty_var = tk.StringVar()
            qty_e = tk.Entry(row_f, textvariable=qty_var, width=10, font=F(11),
                              justify="center", bg=row_f["bg"], relief="solid", bd=1)
            qty_e.pack(side="right", padx=4)

            def _auto_unit(e, iv=item_var, uv=unit_var):
                uv.set(db.get_item_unit(iv.get()))
            item_cb2.bind("<<ComboboxSelected>>", _auto_unit)

            self.item_rows.append((item_var, unit_var, qty_var))

        # ── SIGNATURE ────────────────────────────────────────────────
        sig_card = Card(scroll_inner)
        sig_card.pack(fill="x", padx=16, pady=(0,8))

        sig_row = tk.Frame(sig_card, bg=BG_CARD, padx=14, pady=10)
        sig_row.pack(fill="x")

        # Left: date
        tk.Label(sig_row, text="التاريخ:  _____/_____/_______",
            font=F(11), bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=20)

        # Right: signature space
        sig_box = tk.Frame(sig_row, bg=BG_ROW_ALT, width=200, height=80,
                           relief="solid", bd=1)
        sig_box.pack(side="left", padx=20)
        sig_box.pack_propagate(False)
        tk.Label(sig_box, text="التوقيع",
            font=F(9,False), bg=BG_ROW_ALT, fg=TEXT_LIGHT).pack(expand=True)

        # Notes
        notes_row = tk.Frame(sig_card, bg=BG_CARD, padx=14)
        notes_row.pack(fill="x", pady=(0,10))
        tk.Label(notes_row, text="ملاحظات:", font=F(10,True), bg=BG_CARD).pack(anchor="e")
        self.notes_var = tk.StringVar()
        tk.Entry(notes_row, textvariable=self.notes_var, font=F(11),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1).pack(fill="x")

        # ── ACTION BUTTONS ───────────────────────────────────────────
        btn_card = tk.Frame(scroll_inner, bg=BG_MAIN, padx=16, pady=8)
        btn_card.pack(fill="x")
        ActionBtn(btn_card,"💾  حفظ الشهادة",command=self._save,style="success").pack(side="right",padx=4)
        ActionBtn(btn_card,"📄  تصدير Word",  command=self._export_word, style="primary").pack(side="right",padx=4)
        ActionBtn(btn_card,"🔄  مسح النموذج",command=self._clear,style="neutral").pack(side="right",padx=4)

    def _on_recip_select(self, e):
        name = self.recip_var.get()
        # Get position from staff
        for s in db.get_all_staff():
            if s["name"] == name:
                self.pos_var.set(s["position"] or "")
                break

    def _save(self):
        recip = self.recip_var.get().strip()
        if not recip:
            error_dialog(self,"خطأ","يرجى اختيار المُستلِم"); return
        items = []
        for iv, uv, qv in self.item_rows:
            if iv.get().strip():
                items.append({"item":iv.get(),"unit":uv.get(),"qty":qv.get()})
        if not items:
            error_dialog(self,"خطأ","يرجى إدخال مادة واحدة على الأقل"); return
        db.save_certificate(
            int(self.cert_num_var.get()), recip,
            self.pos_var.get(), self.date_var.get(), json.dumps(items,ensure_ascii=False))
        info_dialog(self,"تم","تم حفظ اشهادة التسلم ✅")
        self.cert_num_var.set(str(db.get_next_cert_number()))
        # Ask if user wants to register items as outgoing stock movements
        self._ask_apply_to_stock(items, recip)

    def _export_word(self):
        """Export certificate using the Word template."""
        recip = self.recip_var.get().strip()
        if not recip:
            error_dialog(self, "خطأ", "يرجى اختيار المُستلِم أولاً"); return
        items = []
        for iv, uv, qv in self.item_rows:
            if iv.get().strip():
                items.append({"item": iv.get(), "unit": uv.get(), "qty": qv.get()})
        if not items:
            error_dialog(self, "خطأ", "يرجى إدخال مادة واحدة على الأقل"); return
        try:
            path = doc_export.export_certificate_docx(
                self.cert_num_var.get(), recip,
                self.pos_var.get(), self.date_var.get(), items
            )
            import subprocess
            subprocess.Popen(["explorer", "/select,", path])
            info_dialog(self, "تصدير Word",
                f"تم تصدير شهادة التسلم إلى:\n{path}")
            # Ask if user wants to register items as outgoing stock
            self._ask_apply_to_stock(items, recip)
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر التصدير:\n{e}")

    def _ask_apply_to_stock(self, items, recipient):
        """
        Prompt to register certificate items as خروج movements in the warehouse.
        """
        answer = messagebox.askyesno(
            "تسجيل في المخزن",
            f"هل تريد تسجيل هذه المواد كعمليات خروج من المخزن\n"
            f"لصالح الأستاذ: {recipient}\n"
            f"(يسمح هذا بتحديث الرصيد تلقائياً)",
            icon="question"
        )
        if not answer:
            return
        date = self.date_var.get()
        saved = 0
        skipped = []
        for it in items:
            item_name = it.get("item", "").strip()
            if not item_name: continue
            try:
                qty = float(it.get("qty", 0) or 0)
            except ValueError:
                qty = 0
            if qty <= 0:
                skipped.append(item_name); continue
            try:
                db.add_movement(
                    item_name=item_name,
                    move_type="خروج",
                    quantity=qty,
                    date=date,
                    beneficiary=recipient,
                    notes=f"شهادة تسلم رقم {self.cert_num_var.get()}"
                )
                saved += 1
            except Exception:
                skipped.append(item_name)
        msg = f"تم تسجيل {saved} عملية خروج في المخزن ✅"
        if skipped:
            msg += f"\nتعذّر تسجيل: {', '.join(skipped)}"
        info_dialog(self, "تم التسجيل", msg)

    def _export_pdf(self):
        """Export certificate as PDF"""
        try:
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os

            out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                f"شهادة_تسليم_{self.cert_num_var.get()}_{datetime.now().strftime('%Y%m%d')}.pdf")

            c = rl_canvas.Canvas(out_path, pagesize=A4)
            w, h = A4

            # Title
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(w/2, h-60, f"شهادة التسلم  رقم: {self.cert_num_var.get()}")
            c.setFont("Helvetica", 12)
            c.drawCentredString(w/2, h-90, "الثانوية الإعدادية ألمدون — جماعة اغيل نومكون")
            c.drawCentredString(w/2, h-110, f"بتاريخ: {self.date_var.get()}")

            y = h - 150
            c.setFont("Helvetica-Bold", 11)
            c.drawRightString(w-40, y, f"المُستلِم: {self.recip_var.get()}")
            y -= 20
            c.drawRightString(w-40, y, f"الصفة: {self.pos_var.get()}")
            y -= 30
            c.drawRightString(w-40, y, "المواد المُسلَّمة:")
            y -= 15

            c.setFont("Helvetica", 10)
            for i,(iv,uv,qv) in enumerate(self.item_rows):
                if iv.get().strip():
                    y -= 22
                    c.drawRightString(w-40, y, f"{i+1}.  {iv.get()}  —  {uv.get()}  —  الكمية: {qv.get()}")

            y -= 60
            c.drawRightString(w-40, y, "التوقيع: ____________________")
            c.drawString(40, y, f"التاريخ: {self.date_var.get()}")

            c.save()
            info_dialog(self, "تصدير PDF",
                f"تم تصدير الشهادة إلى:\n{out_path}")
        except ImportError:
            # Fallback: simple text export
            out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                f"شهادة_{self.cert_num_var.get()}.txt")
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(f"شهادة التسلم رقم: {self.cert_num_var.get()}\n")
                f.write(f"التاريخ: {self.date_var.get()}\n")
                f.write(f"الثانوية الإعدادية ألمدون\n\n")
                f.write(f"المُستلِم: {self.recip_var.get()}\n")
                f.write(f"الصفة: {self.pos_var.get()}\n\n")
                f.write("المواد:\n")
                for i,(iv,uv,qv) in enumerate(self.item_rows):
                    if iv.get(): f.write(f"  {i+1}. {iv.get()} — {uv.get()} — كمية: {qv.get()}\n")
            info_dialog(self,"تصدير",f"تم الحفظ في:\n{out_path}")

    def _clear(self):
        self.recip_var.set(""); self.pos_var.set(""); self.notes_var.set("")
        for iv,uv,qv in self.item_rows:
            iv.set(""); uv.set(""); qv.set("")


# ═══════════════════════════════════════════════════════════════════
#  STOCK CARD PAGE
# ═══════════════════════════════════════════════════════════════════
class StockCardPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        PageHeader(self, "🗃️  بطاقة المخزون — فيش المادة", color=ACCENT).pack(fill="x")

        # Selector
        sel_f = Card(self); sel_f.pack(fill="x", padx=16, pady=10)
        row = tk.Frame(sel_f, bg=BG_CARD, padx=14, pady=10); row.pack(fill="x")

        tk.Label(row, text="الفئة:", font=F(11,True), bg=BG_CARD).pack(side="right", padx=4)
        self.cat_var = tk.StringVar(value=list(db.CATEGORIES.keys())[0])
        cat_cb = ttk.Combobox(row, textvariable=self.cat_var,
            values=list(db.CATEGORIES.keys()), width=20, state="readonly", font=F(11))
        cat_cb.pack(side="right", padx=(0,16))
        cat_cb.bind("<<ComboboxSelected>>", self._on_cat_change)

        tk.Label(row, text="المادة:", font=F(11,True), bg=BG_CARD).pack(side="right", padx=4)
        self.item_var = tk.StringVar()
        self.item_cb = ttk.Combobox(row, textvariable=self.item_var, width=30, font=F(11))
        self.item_cb.pack(side="right", padx=(0,16))
        self.item_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        ActionBtn(row, "🔍  عرض", command=self.refresh, style="primary").pack(side="right", padx=4)
        ActionBtn(row, "🗒️  طباعة البطاقة", command=self._export_fiche, style="success").pack(side="right", padx=4)
        ActionBtn(row, "📦  تصدير جميع البطاقات", command=self._export_all_fiches, style="warning").pack(side="right", padx=4)
        ActionBtn(row, "🗑️  حذف الحركة", command=self._delete_movement, style="danger").pack(side="right", padx=4)
        self._update_item_list()

        # KPI cards
        kpi_f = tk.Frame(self, bg=BG_MAIN, padx=16, pady=8); kpi_f.pack(fill="x")
        self.kpi_cards = {}
        kpis = [
            ("📥","إجمالي الدخول","in",  GREEN,  GREEN_LT),
            ("📤","إجمالي الخروج","out", ORANGE, ORANGE_LT),
            ("📊","الرصيد الحالي","bal", ACCENT, BG_CARD),
            ("🔢","عدد العمليات","ops",  ACCENT3, BG_CARD),
            ("🏷️","الوحدة",      "unit", GOLD,   GOLD_LT),
        ]
        for i,(icon,label,key,fc,bg) in enumerate(kpis):
            card = KPICard(kpi_f, icon, label, "—", color=fc, bg_color=bg)
            card.grid(row=0, column=i, padx=8, sticky="nsew")
            kpi_f.columnconfigure(i, weight=1)
            self.kpi_cards[key] = card

        # Movement history
        tk.Label(self, text="سجل الحركة التفصيلي", font=F(11,True),
                 bg=BG_MAIN, fg=TEXT_DARK, anchor="e", padx=16).pack(fill="x")

        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=16); tbl_f.pack(fill="both", expand=True, pady=(0,10))
        cols  = ("id","date","type","qty","balance","beneficiary","notes")
        hdrs  = ("#","التاريخ","النوع","الكمية","الرصيد","المستفيد / المورد","ملاحظات")
        widths= (0, 110, 80, 80, 90, 200, 200)
        frame, self.tree = make_treeview(tbl_f, cols, hdrs, widths, height=16)
        # Hide the id column
        self.tree.column("id", width=0, minwidth=0, stretch=False)
        self.tree.heading("id", text="")
        frame.pack(fill="both", expand=True)

    def _on_cat_change(self, e):
        self._update_item_list()
        self.item_var.set("")

    def _update_item_list(self):
        items = [r["name"] for r in db.get_items_by_category(self.cat_var.get())]
        self.item_cb["values"] = items

    def refresh(self):
        item = self.item_var.get().strip()
        if not item: return

        in_, out_, bal = db.get_item_balance(item)
        unit = db.get_item_unit(item)
        movements = db.get_movements(item_name=item)

        self.kpi_cards["in"].update_value(int(in_))
        self.kpi_cards["out"].update_value(int(out_))
        self.kpi_cards["bal"].update_value(int(bal))
        self.kpi_cards["ops"].update_value(len(movements))
        self.kpi_cards["unit"].update_value(unit or "—")

        # Color balance card
        bal_color = GREEN if bal > 5 else (ORANGE if bal > 0 else RED)
        self.kpi_cards["bal"].val_label.config(fg=bal_color)

        # Table
        for row in self.tree.get_children(): self.tree.delete(row)
        # Build with running balance
        running_bal = {}
        rb = 0
        chronological_movements = list(reversed(movements))

        for m in chronological_movements:
            if m["type"]=="دخول": rb += m["quantity"]
            else: rb -= m["quantity"]
            running_bal[m["id"]] = rb

        for i, m in enumerate(chronological_movements):
            b = running_bal.get(m["id"], 0)
            tag = "in" if m["type"]=="دخول" else ("out" if b<=0 else ("low" if b<=5 else ("even" if i%2==0 else "odd")))
            self.tree.insert("","end",values=(
                m["id"], m["date"] or "", m["type"], int(m["quantity"]),
                int(b), m["beneficiary"] or "", m["notes"] or ""
            ), tags=(tag,))


    def _delete_movement(self):
        """Delete the selected movement row after confirmation."""
        sel = self.tree.selection()
        if not sel:
            error_dialog(self, "تنبيه", "يرجى تحديد حركة من الجدول أولاً"); return
        vals = self.tree.item(sel[0])["values"]
        mov_id   = int(vals[0])
        mov_date = vals[1]
        mov_type = vals[2]
        mov_qty  = vals[3]
        if not confirm_dialog(self, "⚠️  تأكيد الحذف",
            f"سيتم حذف الحركة:\n{mov_type}  |  الكمية: {mov_qty}  |  التاريخ: {mov_date}\n\n"
            f"هذا الإجراء لا يمكن التراجع عنه. هل تريد المتابعة؟"):
            return
        db.delete_movement(mov_id)
        self.refresh()

    def _export_fiche(self):
        """Export stock card using the Word template fiche_stock.docx."""

        item = self.item_var.get().strip()
        if not item:
            error_dialog(self, "خطأ", "يرجى اختيار المادة أولاً"); return
        unit = db.get_item_unit(item)
        movements = db.get_movements(item_name=item)
        if not movements:
            error_dialog(self, "خطأ", "لا توجد حركات مسجلة لهذه المادة"); return
        try:
            # Build rows with running balance
            rows = []
            bal = 0
            for m in reversed(movements):
                if m["type"] == "دخول":
                    bal += m["quantity"]
                    rows.append({
                        "date":        m["date"] or "",
                        "who":         m["beneficiary"] or "",
                        "type":        m["type"],
                        "quantity":    m["quantity"],
                        "balance":     bal,
                        "notes":       m["notes"] or "",
                    })
                else:
                    bal -= m["quantity"]
                    rows.append({
                        "date":        m["date"] or "",
                        "who":         m["beneficiary"] or "",
                        "type":        m["type"],
                        "quantity":    m["quantity"],
                        "balance":     bal,
                        "notes":       m["notes"] or "",
                    })
            rows.reverse()  # chronological order
            path = doc_export.export_stock_card_docx(item, unit or "", rows)
            import subprocess
            subprocess.Popen(["explorer", "/select,", path])
            info_dialog(self, "تصدير",
                f"تم تصدير بطاقة المخزون إلى:\n{path}")
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر التصدير:\n{e}")

    def _export_all_fiches(self):
        """Export all stock cards into one merged Word document, one fiche per page."""
        ans = messagebox.askyesno(
            "تصدير جميع البطاقات",
            "سيتم دمج بطاقة مخزون لكل مادة لديها حركات في ملف Word واحد،\n"
            "بحيث كل بطاقة في صفحة منفصلة — جاهز للطباعة مباشرة.\n\n"
            "قد يستغرق هذا بضع ثوان. هل تريد المتابعة؟",
            icon="question"
        )
        if not ans:
            return
        try:
            out_path, count, skipped = doc_export.export_all_stock_cards()
            if not out_path:
                error_dialog(self, "لا يوجد", "لا توجد مواد لديها حركات مسجلة."); return
            import subprocess
            subprocess.Popen(["start", "", out_path], shell=True)   # opens in Word
            msg = (f"✅  تم تصدير {count} بطاقة في ملف واحد\n"
                   f"📄  {out_path}\n\n"
                   "افتح الملف ثم اطبعه من Word دفعة واحدة.")
            if skipped:
                msg += f"\n⏭️  {len(skipped)} مادة بدون حركات: تم تخطيها"
            info_dialog(self, "تم التصدير", msg)
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر التصدير:\n{e}")


# ═══════════════════════════════════════════════════════════════════
#  ITEMS MANAGER PAGE
# ═══════════════════════════════════════════════════════════════════
class ItemsManagerPage(tk.Frame):
    """Lets user add, edit, delete items (products) from any category."""
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._selected_id = None
        self._selected_name = None
        self._build()

    def _build(self):
        PageHeader(self, "📝  إدارة المواد والأصناف",
                   subtitle="إضافة أصناف جديدة أو تعديل وحذف الموجودة",
                   color="#4B2E83").pack(fill="x")

        # ── FORM ────────────────────────────────────────────────────
        form = Card(self)
        form.pack(fill="x", padx=16, pady=10)
        tk.Label(form, text="إضافة / تعديل صنف", font=F(11, True),
                 bg=BG_CARD, fg="#4B2E83", anchor="e", padx=14, pady=8).pack(fill="x")

        fields_f = tk.Frame(form, bg=BG_CARD, padx=14)
        fields_f.pack(fill="x", pady=4)

        row = tk.Frame(fields_f, bg=BG_CARD)
        row.pack(fill="x", pady=4)

        # Item name
        tk.Label(row, text="اسم المادة:", font=F(10, True),
                 bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0, 4))
        self.name_var = tk.StringVar()
        tk.Entry(row, textvariable=self.name_var, width=36, font=F(11),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1
                 ).pack(side="right", padx=(0, 16))

        # Unit
        tk.Label(row, text="الوحدة:", font=F(10, True),
                 bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0, 4))
        self.unit_var = tk.StringVar()
        tk.Entry(row, textvariable=self.unit_var, width=14, font=F(11),
                 justify="center", bg=BG_INPUT, relief="solid", bd=1
                 ).pack(side="right", padx=(0, 16))

        # Category
        tk.Label(row, text="الفئة:", font=F(10, True),
                 bg=BG_CARD, fg=TEXT_DARK).pack(side="right", padx=(0, 4))
        self.cat_var = tk.StringVar(value=list(db.CATEGORIES.keys())[0])
        ttk.Combobox(row, textvariable=self.cat_var,
                     values=list(db.CATEGORIES.keys()),
                     width=18, state="readonly", font=F(11)
                     ).pack(side="right", padx=(0, 12))

        btn_f = tk.Frame(form, bg=BG_CARD, padx=14, pady=10)
        btn_f.pack(fill="x")
        ActionBtn(btn_f, "✅  حفظ",         command=self._save,          style="success").pack(side="right", padx=4)
        ActionBtn(btn_f, "✏️  تعديل المحدد", command=self._load_selected, style="primary").pack(side="right", padx=4)
        ActionBtn(btn_f, "🗑️  حذف المحدد",  command=self._delete,        style="danger").pack(side="right", padx=4)
        ActionBtn(btn_f, "🔄  مسح",          command=self._clear,         style="neutral").pack(side="right", padx=4)

        # ── FILTER ──────────────────────────────────────────────────
        flt_f = tk.Frame(self, bg=BG_MAIN, padx=16, pady=4)
        flt_f.pack(fill="x")
        tk.Label(flt_f, text="عرض الفئة:", font=F(10, True), bg=BG_MAIN).pack(side="right", padx=(0, 4))
        self.filter_var = tk.StringVar(value="الكل")
        cats_all = ["الكل"] + list(db.CATEGORIES.keys())
        flt_cb = ttk.Combobox(flt_f, textvariable=self.filter_var,
                               values=cats_all, width=20, state="readonly", font=F(11))
        flt_cb.pack(side="right", padx=4)
        flt_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        # Search
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *a: self.refresh())
        tk.Entry(flt_f, textvariable=self.search_var, width=24, font=F(11),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1).pack(side="right", padx=6)
        tk.Label(flt_f, text="🔍", font=F(12), bg=BG_MAIN).pack(side="right")

        # Counter label
        self.count_lbl = tk.Label(flt_f, text="", font=F(10), bg=BG_MAIN, fg=TEXT_MID)
        self.count_lbl.pack(side="left", padx=8)

        # ── TABLE ────────────────────────────────────────────────────
        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=16)
        tbl_f.pack(fill="both", expand=True, pady=(0, 10))
        cols   = ("id", "name", "unit", "category", "movements")
        hdrs   = ("رقم", "اسم المادة", "الوحدة", "الفئة", "عدد الحركات")
        widths = (50, 280, 120, 160, 110)
        frame, self.tree = make_treeview(tbl_f, cols, hdrs, widths, height=18)
        frame.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        self.refresh()

    def _on_select(self, e):
        sel = self.tree.selection()
        if sel:
            vals = self.tree.item(sel[0])["values"]
            self._selected_id   = vals[0]
            self._selected_name = vals[1]

    def _load_selected(self):
        sel = self.tree.selection()
        if not sel:
            error_dialog(self, "تنبيه", "يرجى تحديد صنف من الجدول أولاً")
            return
        vals = self.tree.item(sel[0])["values"]
        self._selected_id   = vals[0]
        self._selected_name = vals[1]
        self.name_var.set(vals[1])
        self.unit_var.set(vals[2])
        self.cat_var.set(vals[3])

    def _save(self):
        name = self.name_var.get().strip()
        unit = self.unit_var.get().strip()
        cat  = self.cat_var.get().strip()
        if not name:
            error_dialog(self, "خطأ", "يرجى إدخال اسم المادة"); return
        if not unit:
            error_dialog(self, "خطأ", "يرجى إدخال وحدة القياس"); return

        if self._selected_id:
            # Edit mode
            db.update_item(self._selected_id, name, unit, cat)
            info_dialog(self, "تم", f"تم تعديل الصنف ✅")
        else:
            # Add mode
            try:
                db.add_item(name, unit, cat)
                info_dialog(self, "تم", f"تمت إضافة الصنف '{name}' ✅")
            except Exception as ex:
                error_dialog(self, "خطأ", f"لم يتم الحفظ: {ex}"); return
        self._clear()
        self.refresh()

    def _delete(self):
        sel = self.tree.selection()
        if not sel:
            error_dialog(self, "تنبيه", "يرجى تحديد صنف من الجدول أولاً")
            return
        vals = self.tree.item(sel[0])["values"]
        item_id   = vals[0]
        item_name = vals[1]
        mov_count = db.item_has_movements(item_name)

        if mov_count > 0:
            if not confirm_dialog(self, "تحذير",
                f"الصنف '{item_name}' له {mov_count} عملية مسجلة.\n"
                f"هل تريد حذفه فعلاً؟ (لن تُحذف العمليات)"):
                return
        else:
            if not confirm_dialog(self, "تأكيد", f"حذف الصنف '{item_name}'؟"):
                return

        db.delete_item(item_id)
        self._clear()
        self.refresh()

    def _clear(self):
        self.name_var.set("")
        self.unit_var.set("")
        self.cat_var.set(list(db.CATEGORIES.keys())[0])
        self._selected_id   = None
        self._selected_name = None

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        cat_f  = self.filter_var.get()
        search = self.search_var.get().strip().lower()
        items  = db.get_all_items()
        count  = 0
        for i, item in enumerate(items):
            if cat_f != "الكل" and item["category"] != cat_f:
                continue
            if search and search not in item["name"].lower():
                continue
            mvs = db.item_has_movements(item["name"])
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", values=(
                item["id"], item["name"], item["unit"] or "",
                item["category"] or "", mvs
            ), tags=(tag,))
            count += 1
        self.count_lbl.config(text=f"إجمالي الأصناف: {count}")


# ═══════════════════════════════════════════════════════════════════
#  STATISTICS PAGE  —  إعادة تصميم شاملة
# ═══════════════════════════════════════════════════════════════════
class StatisticsPage(tk.Frame):
    # ── Colour palette ──────────────────────────────────────────────
    C = {
        "in":     "#1B998B",   # teal  — دخول
        "out":    "#E84855",   # red   — خروج
        "bal":    "#3A86FF",   # blue  — رصيد
        "ok":     "#06D6A0",   # green — متوفر
        "low":    "#FFB703",   # amber — منخفض
        "empty":  "#EF233C",   # red   — نفد
        "cat1":   "#118AB2",
        "cat2":   "#06D6A0",
        "cat3":   "#FFD166",
        "cat4":   "#EF476F",
        "cat5":   "#8338EC",
        "cat6":   "#FB8500",
        "bg":     "#F4F7FB",
        "card":   "#FFFFFF",
        "dark":   "#1A2740",
        "mid":    "#4A6080",
    }
    PALETTE = ["#118AB2","#06D6A0","#FFD166","#EF476F","#8338EC","#FB8500",
               "#3A86FF","#E84855","#1B998B","#F72585"]

    def __init__(self, parent, **kw):
        super().__init__(parent, bg=self.C["bg"], **kw)
        self._fig = None
        self._canvas_widget = None
        self._build()

    # ────────────────────────────────────────────────────────────────
    def _build(self):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        plt.rcParams.update({
            "font.family":       "DejaVu Sans",
            "axes.spines.top":   False,
            "axes.spines.right": False,
            "axes.grid":         True,
            "grid.alpha":        0.25,
            "axes.facecolor":    "#FAFCFF",
            "figure.facecolor":  "#FAFCFF",
        })

        # ── HEADER ──────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=self.C["dark"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📈  الإحصائيات والتحليلات",
                 font=F(15, True), bg=self.C["dark"], fg="#FFFFFF",
                 anchor="e", padx=20).pack(side="right")
        ActionBtn(hdr, "🔄 تحديث الكل", command=self.refresh,
                  style="neutral").pack(side="left", padx=12)

        # ── KPI CARDS ROW ───────────────────────────────────────────
        kpi_row = tk.Frame(self, bg=self.C["bg"], padx=12, pady=8)
        kpi_row.pack(fill="x")
        self._kpi_frames = {}
        kpi_defs = [
            ("total",     "📦 إجمالي الأصناف", self.C["bal"],   "#FFFFFF"),
            ("available", "✅ متوفر",           self.C["ok"],    "#FFFFFF"),
            ("low",       "⚠️ منخفض",          self.C["low"],   "#1A2740"),
            ("out",       "🔴 نفد المخزون",    self.C["empty"], "#FFFFFF"),
            ("operations","⚡ إجمالي العمليات", self.C["dark"],  "#FFFFFF"),
        ]
        for key, title, bg_c, fg_c in kpi_defs:
            card = tk.Frame(kpi_row, bg=bg_c, padx=16, pady=10,
                            relief="flat", bd=0)
            card.pack(side="right", fill="y", expand=True, padx=4)
            tk.Label(card, text=title, font=F(8, True),
                     bg=bg_c, fg=fg_c, anchor="e").pack(anchor="e")
            val_lbl = tk.Label(card, text="—", font=F(22, True),
                               bg=bg_c, fg=fg_c, anchor="e")
            val_lbl.pack(anchor="e")
            self._kpi_frames[key] = val_lbl

        # ── TAB BUTTONS ─────────────────────────────────────────────
        tab_bar = tk.Frame(self, bg=self.C["bg"], padx=12)
        tab_bar.pack(fill="x", pady=(0, 4))
        self._tab_var = tk.StringVar(value="category")
        tabs = [
            ("🏷️  توزيع الفئات",      "category"),
            ("📅  الاتجاه الشهري",     "monthly"),
            ("🔥  أكثر استهلاكا",      "top_items"),
            ("👤  المستفيدون",         "beneficiaries"),
            ("🚨  تنبيهات المخزون",   "alerts"),
        ]
        self._tab_btns = {}
        for label, key in tabs:
            btn = tk.Button(tab_bar, text=label,
                            font=F(9, True), relief="flat", bd=0,
                            cursor="hand2", padx=10, pady=6,
                            command=lambda k=key: self._switch_tab(k))
            btn.pack(side="right", padx=2)
            self._tab_btns[key] = btn
        self._style_tabs("category")

        # ── CHART FRAME ─────────────────────────────────────────────
        self._chart_outer = tk.Frame(self, bg=self.C["bg"], padx=12)
        self._chart_outer.pack(fill="both", expand=True, pady=(0, 6))

        self._chart_frame = tk.Frame(self._chart_outer, bg=self.C["card"],
                                      relief="flat", bd=0,
                                      highlightbackground="#D0D7E3",
                                      highlightthickness=1)
        self._chart_frame.pack(fill="both", expand=True)

        # ── BOTTOM SUMMARY TABLE ─────────────────────────────────────
        bot = tk.Frame(self, bg=self.C["bg"], padx=12)
        bot.pack(fill="x", pady=(0, 6))
        bot.columnconfigure(0, weight=3); bot.columnconfigure(1, weight=2)

        tk.Label(bot, text="ملخص الفئات", font=F(9, True),
                 bg=self.C["bg"], fg=self.C["dark"], anchor="e"
                 ).grid(row=0, column=0, sticky="e", pady=(0,2), padx=(0,8))
        cf, self._cat_tree = make_treeview(
            bot,
            ("category","in","out","balance","ops"),
            ("الفئة","الدخول","الخروج","الرصيد","العمليات"),
            (130, 80, 80, 80, 80), height=5)
        cf.grid(row=1, column=0, sticky="nsew", padx=(0,8))

        tk.Label(bot, text="أكثر المستفيدين", font=F(9, True),
                 bg=self.C["bg"], fg=self.C["dark"], anchor="e"
                 ).grid(row=0, column=1, sticky="e", pady=(0,2))
        bf, self._ben_tree = make_treeview(
            bot,
            ("name","qty","ops"),
            ("المستفيد","الكمية","العمليات"),
            (160, 80, 80), height=5)
        bf.grid(row=1, column=1, sticky="nsew")

        self.refresh()

    # ────────────────────────────────────────────────────────────────
    def _style_tabs(self, active):
        for key, btn in self._tab_btns.items():
            if key == active:
                btn.config(bg=self.C["dark"], fg="#FFFFFF")
            else:
                btn.config(bg="#D8E4F0", fg=self.C["dark"])

    def _switch_tab(self, key):
        self._tab_var.set(key)
        self._style_tabs(key)
        self._draw_chart()

    # ────────────────────────────────────────────────────────────────
    def _draw_chart(self):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.patches as mpatches

        # Clear previous chart
        for w in self._chart_frame.winfo_children():
            w.destroy()
        if self._fig:
            plt.close(self._fig)
            self._fig = None

        tab = self._tab_var.get()

        try:
            if tab == "category":
                self._chart_category(plt, FigureCanvasTkAgg)
            elif tab == "monthly":
                self._chart_monthly(plt, FigureCanvasTkAgg)
            elif tab == "top_items":
                self._chart_top_items(plt, FigureCanvasTkAgg)
            elif tab == "beneficiaries":
                self._chart_beneficiaries(plt, FigureCanvasTkAgg)
            elif tab == "alerts":
                self._chart_alerts()
        except Exception as e:
            plt.close("all")
            tk.Label(self._chart_frame,
                     text=f"⚠️  تعذّر عرض الرسم البياني\n{e}",
                     font=F(11), bg=self.C["card"], fg=self.C["empty"],
                     anchor="center", justify="center"
                     ).pack(expand=True)

    # ── Chart helpers ────────────────────────────────────────────────
    def _embed_fig(self, fig, plt_mod, FCAgg):
        self._fig = fig
        canvas = FCAgg(fig, master=self._chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt_mod.close(fig)

    def _no_data(self, msg="لا توجد بيانات كافية بعد\nسجّل عمليات أولاً"):
        tk.Label(self._chart_frame, text=f"📊\n{msg}",
                 font=F(13), bg=self.C["card"], fg=self.C["mid"],
                 justify="center").pack(expand=True)

    # ── Tab 1: Category Distribution ─────────────────────────────────
    def _chart_category(self, plt, FCAgg):
        data = db.get_category_stats()
        if not data or not any(r["total_in"] + r["total_out"] > 0 for r in data):
            return self._no_data()

        cats = [r["category"]  for r in data]
        ins  = [r["total_in"]  for r in data]
        outs = [r["total_out"] for r in data]
        bals = [i - o for i, o in zip(ins, outs)]
        safe_bals = [max(0, b) for b in bals]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.8),
                                        facecolor=self.C["card"])
        fig.subplots_adjust(left=0.05, right=0.97, top=0.88, bottom=0.22, wspace=0.3)

        x  = range(len(cats))
        w  = 0.32
        b1 = ax1.bar([i - w/2 for i in x], ins,  w, label="دخول",
                     color=self.C["in"],  alpha=0.9, zorder=3)
        b2 = ax1.bar([i + w/2 for i in x], outs, w, label="خروج",
                     color=self.C["out"], alpha=0.9, zorder=3)
        ax1.bar_label(b1, fmt="%d", fontsize=7, padding=2)
        ax1.bar_label(b2, fmt="%d", fontsize=7, padding=2)
        ax1.set_xticks(list(x))
        ax1.set_xticklabels(cats, rotation=28, ha="right", fontsize=8)
        ax1.set_title("الدخول مقابل الخروج لكل فئة", fontsize=11,
                      fontweight="bold", pad=12)
        ax1.legend(fontsize=9, loc="upper left")

        if sum(safe_bals) > 0:
            wedges, texts, autotexts = ax2.pie(
                safe_bals, labels=None,
                autopct=lambda p: f"{p:.0f}%" if p > 3 else "",
                colors=self.PALETTE[:len(cats)],
                startangle=140, pctdistance=0.78,
                wedgeprops=dict(linewidth=1.5, edgecolor="white"))
            for t in autotexts:
                t.set_fontsize(8)
            ax2.legend(wedges, cats, loc="lower center", fontsize=7,
                       ncol=3, bbox_to_anchor=(0.5, -0.22))
        ax2.set_title("توزيع الرصيد الفعلي", fontsize=11, fontweight="bold", pad=12)

        self._embed_fig(fig, plt, FCAgg)

    # ── Tab 2: Monthly Trend ─────────────────────────────────────────
    def _chart_monthly(self, plt, FCAgg):
        data = db.get_movement_trend(12)
        if not data:
            return self._no_data("لا توجد عمليات محفوظة بعد")

        labels = [r["ym"]       for r in data]
        ins    = [r["total_in"]  for r in data]
        outs   = [r["total_out"] for r in data]
        ops    = [r["ops"]       for r in data]
        x      = range(len(labels))

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 5.2),
                                        facecolor=self.C["card"])
        fig.subplots_adjust(left=0.07, right=0.97, top=0.90, bottom=0.14,
                            hspace=0.55)

        # Grouped bar
        w = 0.38
        ax1.bar([i - w/2 for i in x], ins,  w, label="دخول ↑",
                color=self.C["in"],  alpha=0.9, zorder=3)
        ax1.bar([i + w/2 for i in x], outs, w, label="خروج ↓",
                color=self.C["out"], alpha=0.85, zorder=3)
        ax1.set_xticks(list(x)); ax1.set_xticklabels(labels, fontsize=8, rotation=25)
        ax1.set_title("الكميات الشهرية (دخول / خروج) — آخر 12 شهراً",
                      fontsize=10, fontweight="bold", pad=8)
        ax1.legend(fontsize=9); ax1.set_ylabel("الكمية", fontsize=8)

        # Line chart for ops
        ax2.plot(list(x), ops, marker="o", color=self.C["bal"],
                 linewidth=2.2, markersize=6, zorder=3)
        ax2.fill_between(list(x), ops, alpha=0.12, color=self.C["bal"])
        for xi, yi in zip(x, ops):
            ax2.annotate(str(yi), (xi, yi), textcoords="offset points",
                         xytext=(0, 6), ha="center", fontsize=7,
                         color=self.C["dark"])
        ax2.set_xticks(list(x)); ax2.set_xticklabels(labels, fontsize=8, rotation=25)
        ax2.set_title("عدد العمليات الشهرية", fontsize=10, fontweight="bold", pad=8)
        ax2.set_ylabel("عدد العمليات", fontsize=8)

        self._embed_fig(fig, plt, FCAgg)

    # ── Tab 3: Top Consumed Items ────────────────────────────────────
    def _chart_top_items(self, plt, FCAgg):
        data = db.get_top_consumed_items(12)
        if not data or all(r["total_out"] == 0 for r in data):
            return self._no_data("لم يتم تسجيل أي عمليات خروج بعد")

        items_lbl = [r["item_name"][:22] for r in data]
        outs      = [float(r["total_out"]) for r in data]
        ins       = [float(r["total_in"])  for r in data]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.8),
                                        facecolor=self.C["card"])
        fig.subplots_adjust(left=0.27, right=0.97, top=0.88, bottom=0.1, wspace=0.55)

        colors_out = [self.C["out"] if v > 0 else "#CCCCCC" for v in outs]
        bars = ax1.barh(items_lbl, outs, color=colors_out, alpha=0.9, zorder=3)
        ax1.bar_label(bars, fmt="%d", fontsize=8, padding=3)
        ax1.set_title("🔥 أكثر 12 مادة خروجاً (الكمية)", fontsize=10,
                      fontweight="bold", pad=10)
        ax1.set_xlabel("الكمية المُصرَفة", fontsize=8)
        ax1.tick_params(axis="y", labelsize=8)
        ax1.invert_yaxis()

        colors_in = [self.C["in"] if v > 0 else "#CCCCCC" for v in ins]
        bars2 = ax2.barh(items_lbl, ins, color=colors_in, alpha=0.9, zorder=3)
        ax2.bar_label(bars2, fmt="%d", fontsize=8, padding=3)
        ax2.set_title("📥 أكثر 12 مادة دخولاً (الكمية)", fontsize=10,
                      fontweight="bold", pad=10)
        ax2.set_xlabel("الكمية المُستلَمة", fontsize=8)
        ax2.tick_params(axis="y", labelsize=8)
        ax2.invert_yaxis()

        self._embed_fig(fig, plt, FCAgg)

    # ── Tab 4: Beneficiaries ─────────────────────────────────────────
    def _chart_beneficiaries(self, plt, FCAgg):
        data = db.get_top_beneficiaries(12)
        if not data:
            return self._no_data("لم يتم تسجيل أي خروجات بعد")

        names     = [r["beneficiary"][:20] for r in data]
        qtys      = [float(r["total_qty"])  for r in data]
        ops       = [r["ops"]               for r in data]
        dist_itms = [r["distinct_items"]    for r in data]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.8),
                                        facecolor=self.C["card"])
        fig.subplots_adjust(left=0.25, right=0.97, top=0.88, bottom=0.08, wspace=0.55)

        bar_colors = self.PALETTE[:len(names)]
        bars1 = ax1.barh(names, qtys, color=bar_colors, alpha=0.9, zorder=3)
        ax1.bar_label(bars1, fmt="%d", fontsize=8, padding=3)
        ax1.set_title("👤 إجمالي الكميات المُستلَمة بالشخص",
                      fontsize=10, fontweight="bold", pad=10)
        ax1.set_xlabel("الكمية الإجمالية", fontsize=8)
        ax1.tick_params(axis="y", labelsize=8); ax1.invert_yaxis()

        bars2 = ax2.barh(names, ops, color=bar_colors, alpha=0.75, zorder=3)
        ax2.bar_label(bars2, fmt="%d", fontsize=8, padding=3)
        ax2.set_title("📋 عدد عمليات الاستلام",
                      fontsize=10, fontweight="bold", pad=10)
        ax2.set_xlabel("عدد العمليات", fontsize=8)
        ax2.tick_params(axis="y", labelsize=8); ax2.invert_yaxis()

        self._embed_fig(fig, plt, FCAgg)

    # ── Tab 5: Stock Alerts (no chart — pure Tkinter table) ──────────
    def _chart_alerts(self):
        data = db.get_stock_alerts()

        # Header
        hdr = tk.Frame(self._chart_frame, bg=self.C["card"], pady=6)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🚨  قائمة تنبيهات المخزون",
                 font=F(13, True), bg=self.C["card"],
                 fg=self.C["empty"]).pack(side="right", padx=16)
        total_out = sum(1 for r in data if r["balance"] <= 0)
        total_low = sum(1 for r in data if 0 < r["balance"] <= 5)
        tk.Label(hdr,
                 text=f"نفد: {total_out}  |  منخفض: {total_low}",
                 font=F(10), bg=self.C["card"], fg=self.C["mid"]
                 ).pack(side="right", padx=20)

        if not data:
            tk.Label(self._chart_frame,
                     text="✅  كل المواد في وضع جيد!\nلا توجد تنبيهات.",
                     font=F(14), bg=self.C["card"], fg=self.C["ok"],
                     justify="center").pack(expand=True)
            return

        cols   = ("status", "name", "category", "balance", "last_date")
        hdrs   = ("الحالة", "اسم المادة", "الفئة", "الرصيد", "آخر حركة")
        widths = (80, 260, 130, 80, 110)
        pf, tree = make_treeview(self._chart_frame, cols, hdrs, widths,
                                 height=min(18, len(data)))
        pf.pack(fill="both", expand=True, padx=12, pady=(4, 8))

        tree.tag_configure("out",  background="#FFE5E5", foreground="#CC0000")
        tree.tag_configure("low",  background="#FFF8E1", foreground="#7B5000")

        for r in data:
            bal = int(r["balance"])
            if bal <= 0:
                status = "🔴 نفد المخزون"; tag = "out"
            else:
                status = "🟡 منخفض";       tag = "low"
            tree.insert("", "end", values=(
                status, r["name"], r["category"],
                bal, r["last_date"] or "—"
            ), tags=(tag,))

    # ────────────────────────────────────────────────────────────────
    def refresh(self):
        """Refresh KPI cards, bottom tables, and redraw current chart."""
        # KPIs
        kpi = db.get_kpi_summary()
        for key, lbl in self._kpi_frames.items():
            val = kpi.get(key, 0)
            lbl.config(text=f"{val:,}")

        # Category summary table
        for row in self._cat_tree.get_children():
            self._cat_tree.delete(row)
        for i, r in enumerate(db.get_category_stats()):
            tag = "even" if i % 2 == 0 else "odd"
            self._cat_tree.insert("", "end", values=(
                r["category"], int(r["total_in"]), int(r["total_out"]),
                int(r["total_in"] - r["total_out"]), r["operations"]
            ), tags=(tag,))

        # Beneficiaries table
        for row in self._ben_tree.get_children():
            self._ben_tree.delete(row)
        for i, r in enumerate(db.get_top_beneficiaries(8)):
            tag = "even" if i % 2 == 0 else "odd"
            self._ben_tree.insert("", "end", values=(
                r["beneficiary"], int(r["total_qty"]), r["ops"]
            ), tags=(tag,))

        self._draw_chart()

# ═══════════════════════════════════════════════════════════════════
#  SETTINGS PAGE — إعدادات المؤسسة
# ═══════════════════════════════════════════════════════════════════
class SettingsPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        PageHeader(self, "⚙️  إعدادات المؤسسة",
                   subtitle="معلومات تظهر في الوثائق الرسمية",
                   color="#1A2740").pack(fill="x")

        card = Card(self)
        card.pack(fill="x", padx=20, pady=14)

        tk.Label(card, text="بيانات المؤسسة", font=F(12, True),
                 bg=BG_CARD, fg=ACCENT, anchor="e", padx=14, pady=8).pack(fill="x")

        form = tk.Frame(card, bg=BG_CARD, padx=16); form.pack(fill="x")

        self._vars = {}
        fields = [
            ("اسم المؤسسة",        "school_name"),
            ("الأكاديمية",          "academy"),
            ("المديرية الإقليمية",  "delegation"),
            ("العنوان / المدينة",   "address"),
        ]
        for i, (lbl, key) in enumerate(fields):
            row = tk.Frame(form, bg=BG_CARD); row.pack(fill="x", pady=5)
            tk.Label(row, text=lbl + ":", font=F(11, True),
                     bg=BG_CARD, fg=TEXT_DARK, width=22, anchor="e").pack(side="right")
            v = tk.StringVar()
            self._vars[key] = v
            tk.Entry(row, textvariable=v, font=F(12), justify="right",
                     bg=BG_INPUT, relief="solid", bd=1).pack(
                     side="right", fill="x", expand=True, padx=(0, 8))

        # Logo picker
        logo_row = tk.Frame(form, bg=BG_CARD); logo_row.pack(fill="x", pady=5)
        tk.Label(logo_row, text="شعار المؤسسة:", font=F(11, True),
                 bg=BG_CARD, fg=TEXT_DARK, width=22, anchor="e").pack(side="right")
        self._logo_var = tk.StringVar()
        tk.Entry(logo_row, textvariable=self._logo_var, font=F(10),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1,
                 state="readonly", width=36).pack(side="right", padx=(4, 8))
        ActionBtn(logo_row, "📂 اختيار صورة", command=self._pick_logo,
                  style="primary").pack(side="right")

        # Logo preview label
        self._logo_preview = tk.Label(card, bg=BG_CARD, pady=4)
        self._logo_preview.pack()

        # Buttons
        btn_f = tk.Frame(card, bg=BG_CARD, padx=16, pady=10); btn_f.pack(fill="x")
        ActionBtn(btn_f, "💾  حفظ الإعدادات", command=self._save,
                  style="success").pack(side="right", padx=4)

        tk.Label(self,
                 text="ℹ️  سيظهر الشعار في رأس وثيقة شهادة التسلم",
                 font=F(9), bg=BG_MAIN, fg=TEXT_MID, anchor="e"
                 ).pack(fill="x", padx=20, pady=(4, 0))

        self.refresh()

    def _pick_logo(self):
        path = filedialog.askopenfilename(
            title="اختر صورة الشعار",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp *.gif"),
                       ("All files", "*.*")]
        )
        if path:
            self._logo_var.set(path)
            self._show_preview(path)

    def _show_preview(self, path):
        try:
            from PIL import Image, ImageTk
            img   = Image.open(path).resize((80, 80), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self._logo_preview.config(image=photo, text="")
            self._logo_preview._img = photo   # prevent GC
        except Exception:
            self._logo_preview.config(
                text=f"✅  تم اختيار: {os.path.basename(path)}",
                font=F(9), fg=GREEN)

    def _save(self):
        s = {k: v.get().strip() for k, v in self._vars.items()}
        s["logo_path"] = self._logo_var.get().strip()
        if not s["school_name"]:
            error_dialog(self, "خطأ", "يرجى إدخال اسم المؤسسة على الأقل")
            return
        db.save_settings(
            s["school_name"], s["academy"],
            s["delegation"], s["address"], s["logo_path"]
        )
        info_dialog(self, "تم الحفظ", "✅  تم حفظ إعدادات المؤسسة بنجاح")

    def refresh(self):
        s = db.get_settings()
        for key, v in self._vars.items():
            v.set(s.get(key, ""))
        logo = s.get("logo_path", "")
        self._logo_var.set(logo)
        if logo and os.path.exists(logo):
            self._show_preview(logo)


# ═══════════════════════════════════════════════════════════════════
#  YEAR-END PAGE — إجراءات نهاية السنة
# ═══════════════════════════════════════════════════════════════════
class YearEndPage(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._build()

    def _build(self):
        PageHeader(self, "📅  إجراءات نهاية / بداية السنة",
                   color="#2E4057").pack(fill="x")

        # Year label input
        top = Card(self); top.pack(fill="x", padx=20, pady=12)
        tk.Label(top, text="السنة الدراسية", font=F(12, True),
                 bg=BG_CARD, fg=ACCENT, anchor="e", padx=14, pady=8).pack(fill="x")

        yr_row = tk.Frame(top, bg=BG_CARD, padx=16, pady=8); yr_row.pack(fill="x")
        tk.Label(yr_row, text="رمز السنة (مثال: 2025-2026):",
                 font=F(11, True), bg=BG_CARD, anchor="e").pack(side="right")
        self._year_var = tk.StringVar()
        now = datetime.now()
        y1  = now.year if now.month >= 9 else now.year - 1
        self._year_var.set(f"{y1}-{y1+1}")
        tk.Entry(yr_row, textvariable=self._year_var, font=F(12),
                 justify="right", bg=BG_INPUT, relief="solid", bd=1,
                 width=16).pack(side="right", padx=8)

        # Close year
        cc = Card(self); cc.pack(fill="x", padx=20, pady=6)
        tk.Label(cc,
                 text="📋  إقفال السنة — تصدير كشف المخزون النهائي",
                 font=F(12, True), bg=BG_CARD, fg="#C55A11",
                 anchor="e", padx=14, pady=8).pack(fill="x")
        tk.Label(cc,
                 text="يولّد ملف Excel يحتوي على الرصيد النهائي لكل مادة، ويحفظ لقطة في قاعدة البيانات.",
                 font=F(10), bg=BG_CARD, fg=TEXT_MID, anchor="e", padx=14).pack(fill="x")
        ActionBtn(cc, "📥  إقفال السنة وتصدير الكشف",
                  command=self._close_year, style="warning").pack(
                  anchor="e", padx=14, pady=8)

        # Open new year
        oc = Card(self); oc.pack(fill="x", padx=20, pady=6)
        tk.Label(oc,
                 text="🆕  فتح سنة جديدة — ترحيل الأرصدة",
                 font=F(12, True), bg=BG_CARD, fg=GREEN,
                 anchor="e", padx=14, pady=8).pack(fill="x")
        tk.Label(oc,
                 text="يُسجّل رصيد كل مادة كـ'دخول' (رصيد مرحّل) في بداية السنة الجديدة.",
                 font=F(10), bg=BG_CARD, fg=TEXT_MID, anchor="e", padx=14).pack(fill="x")
        ActionBtn(oc, "🚀  فتح سنة جديدة وترحيل الأرصدة",
                  command=self._open_new_year, style="success").pack(
                  anchor="e", padx=14, pady=8)

        # History
        tk.Label(self, text="سجل الإقفالات السابقة", font=F(11, True),
                 bg=BG_MAIN, fg=TEXT_DARK, anchor="e", padx=20).pack(
                 fill="x", pady=(8, 2))
        tbl_f = tk.Frame(self, bg=BG_MAIN, padx=20)
        tbl_f.pack(fill="both", expand=True)
        frm, self._hist_tree = make_treeview(
            tbl_f,
            ("year", "closed_at"),
            ("السنة الدراسية", "تاريخ الإقفال"),
            (220, 280), height=6)
        frm.pack(fill="both", expand=True)
        self.refresh()

    # ── Handlers ─────────────────────────────────────────────────────
    def _close_year(self):
        year = self._year_var.get().strip()
        if not year:
            error_dialog(self, "خطأ", "يرجى إدخال رمز السنة الدراسية"); return
        if not confirm_dialog(self, "تأكيد الإقفال",
                f"سيتم حفظ الرصيد النهائي لكل المواد لسنة {year}\n"
                f"وتصدير ملف Excel للكشف النهائي.\n\nهل تريد المتابعة؟"):
            return
        try:
            snapshot = db.close_year(year)
            out_path = self._export_year_excel(year, snapshot)
            import subprocess
            subprocess.Popen(["start", "", out_path], shell=True)
            info_dialog(self, "تم الإقفال",
                f"تم إقفال سنة {year} ✅\n"
                f"{len(snapshot)} مادة مُصنَّفة\n"
                f"📂  {out_path}")
            self.refresh()
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر الإقفال:\n{e}")

    def _open_new_year(self):
        year = self._year_var.get().strip()
        if not year:
            error_dialog(self, "خطأ", "يرجى إدخال رمز السنة الدراسية السابقة"); return
        if not confirm_dialog(self, "⚠️  تأكيد فتح السنة الجديدة",
                f"سيتم إنشاء حركات دخول (رصيد مرحّل من {year})\n"
                f"لكل مادة رصيدها أكبر من صفر.\n\n"
                f"تأكد من إقفال السنة السابقة أولاً!\nهل تريد المتابعة؟"):
            return
        try:
            count = db.open_new_year(year)
            info_dialog(self, "تم الترحيل",
                f"✅  تم ترحيل الأرصدة\n"
                f"📦  {count} مادة مُرحَّلة كـ'دخول' في بداية السنة الجديدة")
        except Exception as e:
            error_dialog(self, "خطأ", f"تعذّر الترحيل:\n{e}")

    def _export_year_excel(self, year_label, snapshot):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الكشف النهائي"
        ws.sheet_view.rightToLeft = True

        # Styles
        hdr_font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="1A2740")
        hdr_align = Alignment(horizontal="center", vertical="center")
        c_align   = Alignment(horizontal="right", vertical="center")
        thin      = Side(style="thin", color="D0D7E3")
        bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
        cat_fill  = PatternFill("solid", fgColor="E8EDF5")
        red_fill  = PatternFill("solid", fgColor="FFE5E5")
        amber_fill= PatternFill("solid", fgColor="FFF8E1")

        settings = db.get_settings()
        school   = settings.get("school_name", "المؤسسة") or "المؤسسة"

        # Title
        ws.merge_cells("A1:F1")
        tc = ws["A1"]
        tc.value = f"كشف المخزون النهائي — {school} — {year_label}"
        tc.font  = Font(name="Calibri", bold=True, size=14)
        tc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        # Headers
        hdrs = ["رقم", "اسم المادة", "الفئة", "الرصيد النهائي", "الوحدة", "ملاحظات"]
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = bdr
        ws.row_dimensions[2].height = 22

        prev_cat = None; ri = 3; idx = 0
        for item in snapshot:
            cat = item.get("category", "")
            if cat != prev_cat:
                ws.merge_cells(f"A{ri}:F{ri}")
                cat_c = ws.cell(row=ri, column=1, value=f"▶  {cat}")
                cat_c.font = Font(bold=True, size=11)
                cat_c.fill = cat_fill; cat_c.alignment = c_align
                ws.row_dimensions[ri].height = 18
                ri += 1; prev_cat = cat
            idx += 1
            bal = item.get("balance", 0)
            fill = red_fill if bal <= 0 else (amber_fill if bal <= 5 else None)
            row_data = [idx, item.get("name",""), cat, bal, item.get("unit",""), ""]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = c_align; c.border = bdr
                if fill: c.fill = fill
            ws.row_dimensions[ri].height = 16; ri += 1

        for ci, w in enumerate([8, 42, 22, 16, 14, 20], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        from datetime import datetime as _dt
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
        os.makedirs(exports_dir, exist_ok=True)
        fname = f"إقفال_السنة_{year_label}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        out   = os.path.join(exports_dir, fname)
        wb.save(out)
        return out

    def refresh(self):
        for row in self._hist_tree.get_children():
            self._hist_tree.delete(row)
        for i, r in enumerate(db.get_year_closings()):
            tag = "even" if i % 2 == 0 else "odd"
            self._hist_tree.insert("", "end",
                values=(r["year_label"], r["closed_at"]), tags=(tag,))
